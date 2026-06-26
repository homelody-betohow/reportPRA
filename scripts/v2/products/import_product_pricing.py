from __future__ import annotations

"""
BTH全部SKU明细-*.xlsx → sheet「基础数据维护」 → 表 product_sku_pricing

特点：
- 双行表头（第 1 行一级、第 2 行二级，含合并单元格），数据从第 3 行起，约 5200 条
- 采用固定列索引映射（Excel col 1~51 入库；col 0 仓库 SKU 不入库），并对关键列做表头校验（防列序变动）
- product_sku（第 2 列 / SKU）为业务 UNIQUE KEY：
    * Excel 内同一 product_sku 多行 → 仅保留首条，后续记入跳过明细
    * DB 内 product_sku 已存在 → 按 UPDATABLE_COLS 逐列比较，有差异则 UPDATE
- 新行走 INSERT IGNORE；已存在且有差异的行走显式 UPDATE
- line_hash 仅用于变更检测，幂等键始终是 product_sku（UNIQUE KEY）
- 未指定 --file 时：在共享/本地目录 glob「BTH全部SKU明细-*.xlsx」，按文件名排序后取列表最后一个

用法（在 python/ 目录下）：
    python v2/products/import_product_pricing.py
    python v2/products/import_product_pricing.py --file "\\\\Betohow\\数据报表\\数据库\\BTH全部SKU明细-v2026.04.27.xlsx"
    python v2/products/import_product_pricing.py --sheet 基础数据维护
"""

import argparse
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

_PROD_DIR = Path(__file__).resolve().parent
_V2_DIR = _PROD_DIR.parent
_ORDERS_DIR = _V2_DIR / "orders"
_WR_DIR = _V2_DIR / "warehouse-rent"
for _p in (_PROD_DIR, _V2_DIR, _ORDERS_DIR, _WR_DIR):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

import openpyxl  # type: ignore[import-untyped]

from db import connect, load_db_config
from excel_common import (
    cell_decimal,
    cell_dt,
    cell_int,
    cell_str,
    row_subset_for_line_hash,
    stable_line_hash,
)
from logger import get_logger, setup_stdout_utf8

TABLE = "product_sku_pricing"
SOURCE_TYPE = "Excel"
SHEET_NAME = "基础数据维护"
HEADER_L1_ROW = 1  # 一级表头（1-based）
HEADER_L2_ROW = 2  # 二级表头（1-based）
DATA_START_ROW = 3  # 数据起始行（1-based）

_LOG = get_logger("PRODUCT-PRICING")

# ============================================================
# 固定列索引映射（0-based）
# (列索引, DB字段名, 类型, str截断长度)
# 类型: str / int / dec / date
# ============================================================
COLUMN_MAP: list[tuple[int, str, str, int | None]] = [
    # Excel 第 1 列仍为「仓库 SKU」，仅作人工对照；表 product_sku_pricing 已取消 warehouse_sku 字段，不再入库
    (1,  "product_sku",                "str",  64),   # 业务唯一键（Excel 第 2 列）
    (2,  "cost_price_cny",             "dec",  None),
    (3,  "unit_weight_g",              "dec",  None),
    (4,  "region_spec",                "str",  16),
    # 内箱尺寸
    (5,  "inner_box_l_cm",             "dec",  None),
    (6,  "inner_box_w_cm",             "dec",  None),
    (7,  "inner_box_h_cm",             "dec",  None),
    # 外箱参数
    (8,  "outer_box_l_cm",             "dec",  None),
    (9,  "outer_box_w_cm",             "dec",  None),
    (10, "outer_box_h_cm",             "dec",  None),
    (11, "carton_qty",                 "int",  None),
    (12, "carton_gross_g",             "dec",  None),
    # 头程（RMB/件）
    (13, "first_leg_eu_au_cny",        "dec",  None),
    (14, "first_leg_us_cny",           "dec",  None),
    (15, "first_leg_ca_cny",           "dec",  None),
    (16, "first_leg_jp_cny",           "dec",  None),
    (17, "first_leg_uk_cny",           "dec",  None),
    # 关税含税（RMB/件）
    (18, "duty_eu_cny",                "dec",  None),
    (19, "duty_us_cny",                "dec",  None),
    (20, "duty_ca_au_cny",             "dec",  None),
    (21, "duty_jp_cny",                "dec",  None),
    (22, "duty_uk_cny",                "dec",  None),
    # 关税不含税（RMB/件）
    (23, "duty_eu_notax_cny",          "dec",  None),
    (24, "duty_us_notax_cny",          "dec",  None),
    (25, "duty_ca_au_notax_cny",       "dec",  None),
    (26, "duty_jp_notax_cny",          "dec",  None),
    (27, "duty_uk_notax_cny",          "dec",  None),
    # 采购价格
    (28, "purchase_price_orig_cny",    "dec",  None),
    (29, "supplier_code",              "str",  16),
    (30, "supplier_name",              "str",  128),
    (31, "category",                   "str",  64),
    (32, "ops_mode",                   "str",  16),
    (33, "sales_status",               "str",  32),
    (34, "product_name",               "str",  128),
    (35, "purchase_price_default_cny", "dec",  None),
    (36, "dev_owner",                  "str",  64),
    (37, "supplier_full_name",         "str",  255),
    (38, "product_dev_date",           "date", None),
    (39, "product_launch_date",        "date", None),
    # 佣金点（0~1 小数，如 0.1 表示 10%）
    (40, "commission_pct_regular",     "dec",  None),
    (41, "commission_pct_may_de",      "dec",  None),
    (42, "commission_pct_may_non_de",  "dec",  None),
    (43, "purchase_price_may_cny",     "dec",  None),
    (44, "commission_pct_jun_de",      "dec",  None),
    (45, "commission_pct_jun_non_de",  "dec",  None),
    (46, "purchase_price_jun_cny",     "dec",  None),
    (47, "commission_pct_sep_de",      "dec",  None),
    (48, "commission_pct_sep_non_de",  "dec",  None),
    (49, "purchase_price_sep_cny",     "dec",  None),
    # 单价修改
    (50, "price_modified_cny",         "dec",  None),
    (51, "price_modified_date",        "date", None),
]

# 关键列的预期表头，用于校验 Excel 列序是否发生变动
# 格式：(0-based列索引, 预期L1, 预期L2)；L2 None 表示只校验 L1
_HEADER_CHECKS: list[tuple[int, str, str | None]] = [
    (1,  "SKU",           None),
    (2,  "成本价",         None),
    (3,  "重量（g)",       None),
    (5,  "内箱",          "长（cm)"),
    (8,  "外箱尺寸",       "长（cm)"),
    (13, "头程（RMB）",    "EU/AU"),
    (14, "头程（RMB）",    "US"),
    (17, "头程（RMB）",    "UK"),
    (18, "关税（含税）",   "EU"),
    (22, "关税（含税）",   "UK"),
    (23, "关税（不含税）", "EU"),
    (27, "关税（不含税）", "UK"),
    (28, "原始采购价",     None),
    (29, "供应商代码",     None),
    (31, "品类",           None),
    (32, "运营模式",       None),
    (33, "产品销售状态",   None),
    (38, "产品开发时间",   None),
    (39, "产品上架时间",   None),
]

# 写入表时使用的全部列（除 id/created_at/updated_at）
INSERT_COLS: list[str] = ["line_hash"] + [col for _, col, _, _ in COLUMN_MAP] + ["source_type"]

# 已存在 product_sku 时允许被 Excel 新值覆盖的列
UPDATABLE_COLS: tuple[str, ...] = tuple(c for c in INSERT_COLS if c != "product_sku")

# 参与 line_hash 的核心业务字段（谨慎修改，改后历史 hash 与库内不符）
LINE_HASH_KEYS: tuple[str, ...] = (
    "product_sku",
    "cost_price_cny",
    "unit_weight_g",
    "carton_qty",
    "carton_gross_g",
    "first_leg_eu_au_cny",
    "first_leg_us_cny",
    "first_leg_ca_cny",
    "first_leg_jp_cny",
    "first_leg_uk_cny",
    "duty_eu_cny",
    "duty_us_cny",
    "duty_ca_au_cny",
    "duty_jp_cny",
    "duty_uk_cny",
    "duty_eu_notax_cny",
    "duty_us_notax_cny",
    "duty_ca_au_notax_cny",
    "duty_jp_notax_cny",
    "duty_uk_notax_cny",
    "purchase_price_orig_cny",
    "purchase_price_default_cny",
    "purchase_price_may_cny",
    "purchase_price_jun_cny",
    "purchase_price_sep_cny",
    "supplier_code",
    "ops_mode",
    "sales_status",
    "source_type",
)


# ============================================================
# 类型转换
# ============================================================
def _cell_date(v: Any) -> date | None:
    """datetime / date / 字符串 → date（仅日期部分）。"""
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    dt = cell_dt(v)
    return dt.date() if dt else None


def _convert(v: Any, kind: str, max_len: int | None) -> Any:
    if kind == "int":
        return cell_int(v)
    if kind == "dec":
        return cell_decimal(v)
    if kind == "date":
        return _cell_date(v)
    return cell_str(v, max_len=max_len)


# ============================================================
# Excel 读取 + 表头校验
# ============================================================
def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", "").replace("\r", "").strip()


def _validate_headers(ws) -> None:
    """
    按 _HEADER_CHECKS 校验关键列的表头是否符合预期。
    任意一列不匹配则打印警告（不中断，允许轻微表头差异）。
    """
    rows = list(ws.iter_rows(min_row=HEADER_L1_ROW, max_row=HEADER_L2_ROW, values_only=True))
    if len(rows) < 2:
        _LOG.warn(f"sheet「{ws.title}」表头行不足，跳过校验")
        return
    row_l1, row_l2 = rows[0], rows[1]
    mismatches: list[str] = []
    for col_idx, exp_l1, exp_l2 in _HEADER_CHECKS:
        actual_l1 = _norm(row_l1[col_idx] if col_idx < len(row_l1) else None)
        actual_l2 = _norm(row_l2[col_idx] if col_idx < len(row_l2) else None)
        if exp_l1 and actual_l1 != exp_l1:
            mismatches.append(f"col[{col_idx}] L1 期望={exp_l1!r} 实际={actual_l1!r}")
        if exp_l2 and actual_l2 != exp_l2:
            mismatches.append(f"col[{col_idx}] L2 期望={exp_l2!r} 实际={actual_l2!r}")
    if mismatches:
        _LOG.warn(f"表头校验发现 {len(mismatches)} 处不符（列序可能已变动，请核查）：")
        for m in mismatches:
            _LOG.warn(f"  {m}")
    else:
        _LOG.info("表头校验通过：关键列均在预期位置")


def _row_to_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for col_idx, db_col, kind, max_len in COLUMN_MAP:
        v = row[col_idx] if col_idx < len(row) else None
        out[db_col] = _convert(v, kind, max_len)
    out["source_type"] = SOURCE_TYPE
    return out


# ============================================================
# DB 操作
# ============================================================
def _existing_skus(conn, skus: list[str]) -> set[str]:
    skus = [s for s in skus if s]
    if not skus:
        return set()
    cur = conn.cursor()
    found: set[str] = set()
    chunk = 500
    for i in range(0, len(skus), chunk):
        batch = skus[i : i + chunk]
        placeholders = ",".join(["%s"] * len(batch))
        cur.execute(
            f"SELECT `product_sku` FROM `{TABLE}` WHERE `product_sku` IN ({placeholders})",
            batch,
        )
        for (sku,) in cur.fetchall():
            found.add(sku)
    cur.close()
    return found


def _fetch_rows_for_update(conn, skus: list[str]) -> dict[str, dict[str, Any]]:
    skus = [s for s in skus if s]
    if not skus:
        return {}
    cols_sql = ", ".join(f"`{c}`" for c in UPDATABLE_COLS)
    cur = conn.cursor()
    out: dict[str, dict[str, Any]] = {}
    chunk = 500
    for i in range(0, len(skus), chunk):
        batch = skus[i : i + chunk]
        placeholders = ",".join(["%s"] * len(batch))
        sql = f"SELECT `product_sku`, {cols_sql} FROM `{TABLE}` WHERE `product_sku` IN ({placeholders})"
        cur.execute(sql, batch)
        col_names = ["product_sku"] + list(UPDATABLE_COLS)
        for row in cur.fetchall():
            sku = row[0]
            out[sku] = {col_names[j]: row[j] for j in range(1, len(col_names))}
    cur.close()
    return out


def _cmp_cell(a: Any, b: Any) -> bool:
    if a is None and b is None:
        return True
    if isinstance(a, str) and a.strip() == "":
        a = None  # type: ignore[assignment]
    if isinstance(b, str) and b.strip() == "":
        b = None  # type: ignore[assignment]
    if a is None or b is None:
        return a is None and b is None
    if isinstance(a, (Decimal, float)) or isinstance(b, (Decimal, float)):
        try:
            return Decimal(str(a)) == Decimal(str(b))
        except Exception:
            return a == b
    if isinstance(a, date) or isinstance(b, date):
        try:
            da = a.date() if isinstance(a, datetime) else a
            db_ = b.date() if isinstance(b, datetime) else b
            return da == db_
        except Exception:
            return a == b
    if isinstance(a, (int, float)) or isinstance(b, (int, float)):
        try:
            return int(a) == int(b)
        except Exception:
            return a == b
    return a == b


def _needs_update(old: dict[str, Any], new: dict[str, Any]) -> bool:
    for c in UPDATABLE_COLS:
        if not _cmp_cell(old.get(c), new.get(c)):
            return True
    return False


def _insert_rows(conn, rows: list[tuple[Any, ...]]) -> int:
    if not rows:
        return 0
    cols_sql = ", ".join(f"`{c}`" for c in INSERT_COLS)
    placeholders = ", ".join(["%s"] * len(INSERT_COLS))
    sql = f"INSERT IGNORE INTO `{TABLE}` ({cols_sql}) VALUES ({placeholders})"
    cur = conn.cursor()
    affected = 0
    chunk = 300
    for i in range(0, len(rows), chunk):
        batch = rows[i : i + chunk]
        cur.executemany(sql, batch)
        affected += max(cur.rowcount or 0, 0)
    cur.close()
    return affected


def _update_rows(conn, params: list[tuple[Any, ...]]) -> int:
    if not params:
        return 0
    set_clause = ", ".join(f"`{c}`=%s" for c in UPDATABLE_COLS)
    sql = f"UPDATE `{TABLE}` SET {set_clause} WHERE `product_sku`=%s"
    cur = conn.cursor()
    affected = 0
    chunk = 300
    for i in range(0, len(params), chunk):
        batch = params[i : i + chunk]
        cur.executemany(sql, batch)
        rc = cur.rowcount
        affected += len(batch) if rc is None or rc < 0 else rc
    cur.close()
    return max(affected, 0)


# ============================================================
# 主导入流程
# ============================================================
def import_file(conn, xlsx: Path, sheet_name: str = SHEET_NAME) -> dict[str, int]:
    """
    Returns:
        {"excel_rows": ..., "inserted": ..., "updated": ..., "skipped": ...}
    """
    _LOG.warn(f"读取 Excel：{xlsx}  sheet={sheet_name!r}")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"未找到 sheet「{sheet_name}」，可用：{wb.sheetnames}")
    ws = wb[sheet_name]

    _validate_headers(ws)

    # ===== 第一遍：读取并转换所有数据行，Excel 内部去重 =====
    seen_sku: set[str] = set()
    candidates: list[dict[str, Any]] = []
    skipped = 0
    excel_rows = 0

    for ridx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        # 判空行：product_sku（col 1）为空则跳过
        if not row or row[1] is None or (isinstance(row[1], str) and not row[1].strip()):
            continue
        excel_rows += 1
        d = _row_to_dict(row)
        sku = d.get("product_sku")
        if not sku:
            skipped += 1
            _LOG.warn(f"第 {ridx} 行缺 product_sku（SKU列为空），已跳过")
            continue
        if sku in seen_sku:
            skipped += 1
            _LOG.warn(f"第 {ridx} 行 product_sku={sku!r} 在 Excel 内重复，保留首次出现，后续跳过")
            continue
        seen_sku.add(sku)
        d["_excel_row"] = ridx
        candidates.append(d)

    _LOG.info(
        f"扫描完成：Excel数据行={excel_rows}  候选行={len(candidates)}  跳过={skipped}"
    )

    # ===== 第二遍：批量查 DB =====
    existed = _existing_skus(conn, [d["product_sku"] for d in candidates])
    _LOG.info(f"DB 内已存在 product_sku：{len(existed)} 个")
    existing_rows = _fetch_rows_for_update(conn, list(existed))

    # ===== 第三遍：分拣 INSERT / UPDATE =====
    rows_to_insert: list[tuple[Any, ...]] = []
    rows_to_update: list[tuple[Any, ...]] = []

    for d in candidates:
        sku = d["product_sku"]
        d["line_hash"] = stable_line_hash(row_subset_for_line_hash(d, LINE_HASH_KEYS))
        if sku in existed:
            old = existing_rows.get(sku)
            if old is None or _needs_update(old, d):
                vals = tuple(d[c] for c in UPDATABLE_COLS) + (sku,)
                rows_to_update.append(vals)
        else:
            rows_to_insert.append(tuple(d[c] for c in INSERT_COLS))

    _LOG.info(f"准备 INSERT IGNORE：{len(rows_to_insert)} 行  UPDATE：{len(rows_to_update)} 行")
    inserted = _insert_rows(conn, rows_to_insert)
    updated = _update_rows(conn, rows_to_update)
    _LOG.info(
        f"导入完成：Excel数据行={excel_rows}  插入={inserted}  更新={updated}  跳过={skipped}"
    )
    return {"excel_rows": excel_rows, "inserted": inserted, "updated": updated, "skipped": skipped}


# ============================================================
# 默认文件路径
# ============================================================
def default_excel_path() -> Path:
    """
    自动在网络共享目录中查找 BTH全部SKU明细-*.xlsx，按文件名排序后仅取列表最后一个。
    若网络路径不可访问，退化到本地 python/excel/base/ 目录。
    """
    net_dir = Path(r"\\Betohow\数据报表\数据库")
    if net_dir.is_dir():
        candidates = sorted(net_dir.glob("BTH全部SKU明细-*.xlsx"), reverse=True)
        if candidates:
            return candidates[-1]
    local_dir = Path(__file__).resolve().parents[3] / "excel" / "base"
    candidates = sorted(local_dir.glob("BTH全部SKU明细-*.xlsx"), reverse=True)
    if candidates:
        return candidates[-1]
    return net_dir / "BTH全部SKU明细.xlsx"


# ============================================================
# CLI 入口
# ============================================================
def main() -> int:
    setup_stdout_utf8()
    ap = argparse.ArgumentParser(
        description="导入 BTH全部SKU明细-*.xlsx 基础数据维护 -> product_sku_pricing"
    )
    ap.add_argument(
        "--file",
        type=Path,
        default=None,
        help="指定 xlsx 路径；默认在共享/本地目录 glob 后取排序列表的最后一个文件",
    )
    ap.add_argument("--sheet", default=SHEET_NAME, help=f"sheet 名称，默认「{SHEET_NAME}」")
    args = ap.parse_args()

    xlsx = args.file or default_excel_path()
    if not xlsx.is_file():
        _LOG.error(f"Excel 文件不存在或无法访问：{xlsx}")
        return 2

    _LOG.info(f"任务：{xlsx.name} → {TABLE}")

    cfg = load_db_config()
    _LOG.info(f"连接数据库：host={cfg.host} port={cfg.port} database={cfg.database} user={cfg.user}")
    conn = connect(cfg)
    try:
        stats = import_file(conn, xlsx, sheet_name=args.sheet)
        conn.commit()
        _LOG.info(
            f"已提交事务：Excel行={stats['excel_rows']}  "
            f"插入={stats['inserted']}  更新={stats['updated']}  跳过={stats['skipped']}"
        )
        return 0
    except Exception:
        conn.rollback()
        _LOG.error("发生异常，已回滚事务")
        raise
    finally:
        conn.close()
        _LOG.info("数据库连接已关闭")


if __name__ == "__main__":
    raise SystemExit(main())
