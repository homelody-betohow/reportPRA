from __future__ import annotations

"""
产品信息库导入：python/excel/base/产品信息库.xlsx → 表 product_sku

特点：
- 处理 Excel 多行表头（第 5 行一级表头 + 第 6 行二级表头，含合并单元格）
- product_sku（产品编码）为 UNIQUE KEY：
    * Excel 内同一 product_sku 多行 → 仅保留首条，后续记录到「跳过明细」
    * DB 内 product_sku 已存在 → 按 UPDATABLE_COLS 比较，有变化则 UPDATE（不记入跳过）
- 跳过明细完成后，若 .env 配置了 SMTP_*，发送 HTML 邮件通知
- 新行仍走 INSERT IGNORE；已存在行走显式 UPDATE（并发下 INSERT 仍 IGNORE 兜底）

用法（在 python/ 目录下）：
    python v2/products/import_product_master.py
    python v2/products/import_product_master.py --file path/to/产品信息库.xlsx
    python v2/products/import_product_master.py --no-mail   # 即便有跳过也不发邮件
    python v2/products/import_product_master.py --always-mail  # 无跳过也发一封"成功"邮件
"""

import argparse
import os
import smtplib
import sys
from dataclasses import dataclass, field
from datetime import datetime
from email.message import EmailMessage
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

_PROD_DIR = Path(__file__).resolve().parent
_V2_DIR = _PROD_DIR.parent
_ORDERS_DIR = _V2_DIR / "orders"
_WR_DIR = _V2_DIR / "warehouse-rent"
for _p in (_PROD_DIR, _V2_DIR, _ORDERS_DIR, _WR_DIR):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

import openpyxl  # type: ignore[import-untyped]

from db import connect, load_db_config
from excel_common import (
    cell_decimal,
    cell_int,
    cell_str,
    row_subset_for_line_hash,
    stable_line_hash,
    v2_root,
)
from logger import get_logger, setup_stdout_utf8

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None  # type: ignore[assignment]


TABLE = "product_sku"
SOURCE_TYPE = "Excel"
SHEET_NAME = "产品信息表"
HEADER_L1_ROW = 5  # 一级表头
HEADER_L2_ROW = 6  # 二级表头
DATA_START_ROW = 7

_LOG = get_logger("PRODUCT-SKU")


# ============================================================
# 字段映射：(L1 表头, L2 表头, DB 列名, 类型, str 截断长度)
# 类型：str / int / dec
# L2 为 None 表示该列只有一级表头（无合并）
# ============================================================
COLUMN_MAP: list[tuple[str, str | None, str, str, int | None]] = [
    # 业务标识
    ("产品编码", None, "product_sku", "str", 64),
    ("商品ID", None, "product_uid", "str", 64),
    ("仓库识别码", None, "warehouse_ref", "str", 64),
    # 分类
    ("二级分类", None, "category_lv2", "str", 64),
    ("三级分类", None, "category_lv3", "str", 64),
    # 含换行符的合并表头单元格：原文为「产品类别代码\n（根据命名规则）」
    ("产品类别代码（根据命名规则）", None, "category_code", "str", 8),
    # 供应商
    ("供应商", None, "supplier_name", "str", 128),
    # 基础属性
    ("单位", None, "product_unit", "str", 16),
    ("产品颜色", None, "product_color", "str", 32),
    # 生命周期/核算
    ("AMZ新老品", None, "amz_lifecycle", "str", 16),
    ("本土平台新老品", None, "local_lifecycle", "str", 16),
    ("核算分类", None, "accounting_class", "str", 32),
    # 采购参数
    ("MOQ", None, "purchase_moq", "int", None),
    ("采购交期", None, "purchase_lead_days", "int", None),
    ("箱规", None, "carton_qty", "int", None),
    ("成本价", None, "cost_price_cny", "dec", None),
    # 重量
    ("重量（g)", None, "unit_weight_g", "dec", None),
    # 内箱尺寸（合并表头）
    ("内箱", "长（cm)", "inner_box_l_cm", "dec", None),
    ("内箱", "宽（cm)", "inner_box_w_cm", "dec", None),
    ("内箱", "高（cm)", "inner_box_h_cm", "dec", None),
    # 外箱尺寸（合并表头）
    ("外箱尺寸", "长（cm)", "outer_box_l_cm", "dec", None),
    ("外箱尺寸", "宽（cm)", "outer_box_w_cm", "dec", None),
    ("外箱尺寸", "高（cm)", "outer_box_h_cm", "dec", None),
    ("外箱尺寸", "箱规毛重（g）", "carton_gross_g", "dec", None),
    # 头程（RMB）
    ("头程（RMB）", "EU/AU", "first_leg_eu_au_cny", "dec", None),
    ("头程（RMB）", "US", "first_leg_us_cny", "dec", None),
    ("头程（RMB）", "UK", "first_leg_uk_cny", "dec", None),
    # 关税（RMB）
    ("关税（RMB）", "EU", "duty_eu_cny", "dec", None),
    ("关税（RMB）", "US", "duty_us_cny", "dec", None),
    ("关税（RMB）", "UK", "duty_uk_cny", "dec", None),
]

# 写入 product_sku 表时使用的全部列（除 id/created_at/updated_at）
INSERT_COLS: list[str] = ["line_hash"] + [c for _, _, c, _, _ in COLUMN_MAP] + ["source_type"]

# 已存在 product_sku 时允许用 Excel 新值覆盖的列（不含主键 product_sku）
# 若需「冻结」某些字段（仅首录、永不覆盖），从元组中移除对应列名即可
UPDATABLE_COLS: tuple[str, ...] = tuple(c for c in INSERT_COLS if c != "product_sku")

# line_hash 参与字段：除主键 product_sku 外的核心识别 + 价格关键值
# 改这里会导致历史 line_hash 与库内不一致；用于变更检测，不影响 UPSERT 幂等性（幂等键是 product_sku）
LINE_HASH_KEYS: tuple[str, ...] = (
    "product_sku",
    "product_uid",
    "warehouse_ref",
    "category_lv2",
    "category_lv3",
    "supplier_name",
    "product_color",
    "amz_lifecycle",
    "local_lifecycle",
    "accounting_class",
    "cost_price_cny",
    "carton_qty",
    "unit_weight_g",
    "carton_gross_g",
    "inner_box_l_cm",
    "inner_box_w_cm",
    "inner_box_h_cm",
    "outer_box_l_cm",
    "outer_box_w_cm",
    "outer_box_h_cm",
    "first_leg_eu_au_cny",
    "first_leg_us_cny",
    "first_leg_uk_cny",
    "duty_eu_cny",
    "duty_us_cny",
    "duty_uk_cny",
    "source_type"
)


# ============================================================
# 跳过记录
# ============================================================
@dataclass
class SkipRecord:
    excel_row: int
    product_sku: str
    product_uid: str | None
    supplier_name: str | None
    product_color: str | None
    reason: str  # "Excel内重复" / "DB内已存在" / "缺产品编码"


@dataclass
class ImportStats:
    excel_rows: int = 0           # Excel 数据行数（去掉空行）
    inserted: int = 0             # 实际插入到 DB 的行数
    updated: int = 0            # 已存在 SKU 且允许列有变化时 UPDATE 的行数
    skipped_records: list[SkipRecord] = field(default_factory=list)
    started_at: str = ""
    ended_at: str = ""
    file_name: str = ""

    @property
    def skipped(self) -> int:
        return len(self.skipped_records)

    @property
    def skip_reasons(self) -> dict[str, int]:
        """按原因聚合的跳过计数，便于日志输出和邮件决策。"""
        out: dict[str, int] = {}
        for r in self.skipped_records:
            out[r.reason] = out.get(r.reason, 0) + 1
        return out

    @property
    def has_only_db_existed_skips(self) -> bool:
        """兼容旧版「DB内已存在」跳过原因；当前重复 SKU 会尝试 UPDATE，一般不再产生该跳过。"""
        if not self.skipped_records:
            return False
        return all(r.reason == "DB内已存在" for r in self.skipped_records)


# ============================================================
# Excel 读取（处理多行表头 + 合并单元格）
# ============================================================
def _norm(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", "").replace("\r", "").strip()


def _build_col_index(ws) -> dict[tuple[str, str | None], int]:
    """
    扫描第 5/6 行表头，返回 (L1, L2) → 0-based 列索引 的字典。
    合并单元格在 openpyxl read_only 下只有左上角有值，这里手动「向后补全」L1：
    若 L2 非空但 L1 为空，则继承上一个非空 L1。
    """
    rows = list(ws.iter_rows(min_row=HEADER_L1_ROW, max_row=HEADER_L2_ROW, values_only=True))
    if len(rows) < 2:
        raise RuntimeError(f"sheet「{ws.title}」缺少表头：需要第 {HEADER_L1_ROW}/{HEADER_L2_ROW} 行")
    row_l1, row_l2 = rows[0], rows[1]

    out: dict[tuple[str, str | None], int] = {}
    last_l1 = ""
    n = max(len(row_l1), len(row_l2))
    for i in range(n):
        h1 = _norm(row_l1[i] if i < len(row_l1) else None)
        h2 = _norm(row_l2[i] if i < len(row_l2) else None)
        if h1:
            last_l1 = h1
        # 决定有效 L1：仅当 L2 有值或本列 L1 自身有值时，才认为是「合并单元格归属」
        eff_l1 = h1 if h1 else (last_l1 if h2 else "")
        eff_l2: str | None = h2 if h2 else None
        if not eff_l1 and not eff_l2:
            continue
        # 同一 (L1, L2) 出现重复时，保留首个（极少发生）
        key = (eff_l1, eff_l2)
        if key not in out:
            out[key] = i
    return out


def _resolve_columns(col_index: dict[tuple[str, str | None], int]) -> dict[str, int]:
    """COLUMN_MAP 的每一项 → Excel 0-based 列索引；缺失列报警但允许继续（写 None）。"""
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for l1, l2, db_col, _, _ in COLUMN_MAP:
        idx = col_index.get((l1, l2))
        if idx is None:
            missing.append(f"{l1}/{l2}" if l2 else l1)
        else:
            resolved[db_col] = idx
    if missing:
        _LOG.warn(f"Excel 表头缺失列（将写 NULL）：{missing}")
    return resolved


def _convert(v: Any, kind: str, max_len: int | None) -> Any:
    if kind == "int":
        return cell_int(v)
    if kind == "dec":
        return cell_decimal(v)
    return cell_str(v, max_len=max_len)


def _row_to_dict(row: tuple[Any, ...], col_idx: dict[str, int]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for l1, l2, db_col, kind, max_len in COLUMN_MAP:
        i = col_idx.get(db_col)
        v = row[i] if i is not None and i < len(row) else None
        out[db_col] = _convert(v, kind, max_len)
    out["source_type"] = SOURCE_TYPE
    return out


# ============================================================
# DB 操作
# ============================================================
def _existing_product_skus(conn, skus: Iterable[str]) -> set[str]:
    """批量查询已存在的 product_sku，返回集合。"""
    skus = [s for s in skus if s]
    if not skus:
        return set()
    cur = conn.cursor()
    found: set[str] = set()
    chunk = 500
    for i in range(0, len(skus), chunk):
        batch = skus[i : i + chunk]
        placeholders = ",".join(["%s"] * len(batch))
        cur.execute(f"SELECT `product_sku` FROM `{TABLE}` WHERE `product_sku` IN ({placeholders})", batch)
        for (sku,) in cur.fetchall():
            found.add(sku)
    cur.close()
    return found


def _fetch_rows_for_update(conn, skus: list[str]) -> dict[str, dict[str, Any]]:
    """按 product_sku 拉取 UPDATABLE_COLS 当前值，用于与 Excel 行比较。"""
    skus = [s for s in skus if s]
    if not skus:
        return {}
    cols_sql = ", ".join(f"`{c}`" for c in UPDATABLE_COLS)
    col_names = ["product_sku"] + list(UPDATABLE_COLS)
    cur = conn.cursor()
    out: dict[str, dict[str, Any]] = {}
    chunk = 500
    for i in range(0, len(skus), chunk):
        batch = skus[i : i + chunk]
        placeholders = ",".join(["%s"] * len(batch))
        sql = f"SELECT `product_sku`, {cols_sql} FROM `{TABLE}` WHERE `product_sku` IN ({placeholders})"
        cur.execute(sql, batch)
        for row in cur.fetchall():
            sku = row[0]
            out[sku] = {col_names[j]: row[j] for j in range(1, len(col_names))}
    cur.close()
    return out


def _cmp_cell(a: Any, b: Any) -> bool:
    """比较 Excel 转换值与 DB 读回值是否视为相同。"""
    if a is None and b is None:
        return True
    if isinstance(a, str) and a.strip() == "":
        a = None  # type: ignore[assignment]
    if isinstance(b, str) and b.strip() == "":
        b = None  # type: ignore[assignment]
    if a is None or b is None:
        return a is None and b is None
    if isinstance(a, Decimal) or isinstance(b, Decimal):
        try:
            return Decimal(str(a)) == Decimal(str(b))
        except Exception:
            return a == b
    if isinstance(a, (int, float)) or isinstance(b, (int, float)):
        if isinstance(a, bool) or isinstance(b, bool):
            return a == b
        try:
            return int(a) == int(b)
        except Exception:
            return a == b
    return a == b


def _needs_update(old: dict[str, Any], new: dict[str, Any]) -> bool:
    for c in UPDATABLE_COLS:
        if not _cmp_cell(old.get(c), new.get(c)):
            return True
    return False


def _insert_rows(conn, rows: list[tuple[Any, ...]]) -> int:
    """
    INSERT IGNORE INTO product_sku (...) VALUES (...);
    使用 IGNORE 兜底：若并发场景下又有别的进程抢先插入了同 product_sku，
    本次也不会抛异常，安静跳过。
    """
    if not rows:
        return 0
    cols_sql = ", ".join(f"`{c}`" for c in INSERT_COLS)
    placeholders = ", ".join(["%s"] * len(INSERT_COLS))
    sql = f"INSERT IGNORE INTO `{TABLE}` ({cols_sql}) VALUES ({placeholders})"
    cur = conn.cursor()
    affected = 0
    chunk = 300
    for i in range(0, len(rows), chunk):
        batch = rows[i : i + chunk]
        cur.executemany(sql, batch)
        affected += cur.rowcount
    cur.close()
    return max(affected, 0)


def _update_rows(conn, params: list[tuple[Any, ...]]) -> int:
    """
    按 product_sku 更新 UPDATABLE_COLS。每行参数：按 UPDATABLE_COLS 顺序的新值，最后一列为 WHERE 的 product_sku。
    """
    if not params:
        return 0
    set_clause = ", ".join(f"`{c}`=%s" for c in UPDATABLE_COLS)
    sql = f"UPDATE `{TABLE}` SET {set_clause} WHERE `product_sku`=%s"
    cur = conn.cursor()
    affected = 0
    chunk = 300
    for i in range(0, len(params), chunk):
        batch = params[i : i + chunk]
        cur.executemany(sql, batch)
        rc = cur.rowcount
        affected += len(batch) if rc is None or rc < 0 else rc
    cur.close()
    return max(affected, 0)


# ============================================================
# 邮件通知（复用 .env 中 SMTP_* 配置）
# ============================================================
def _load_env_files() -> None:
    if not load_dotenv:
        return
    here = Path(__file__).resolve().parent
    for env_path in (
        here.parent.parent / ".env",        # python/.env
        here.parent / ".env",               # python/v2/.env
        here.parent.parent.parent / ".env",  # 仓库根 .env
    ):
        if env_path.is_file():
            load_dotenv(env_path)
            return


def _env_bool(name: str, default: bool) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    return v.strip().lower() in {"1", "true", "yes", "y", "on"}


def _esc(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def _build_skip_html(stats: ImportStats) -> str:
    rows_html = "\n".join(
        f"""
        <tr>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;text-align:right;color:#6b7280;">{r.excel_row}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;font-family:Menlo,Consolas,monospace;color:#111827;">{_esc(r.product_sku)}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#374151;">{_esc(r.product_uid or '')}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#374151;">{_esc(r.supplier_name or '')}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#374151;">{_esc(r.product_color or '')}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#b45309;">{_esc(r.reason)}</td>
        </tr>
        """.strip()
        for r in stats.skipped_records
    )
    badge = (
        f'<span style="display:inline-block;padding:4px 10px;border-radius:999px;'
        f'background:#fef3c7;color:#92400e;font-weight:700;font-size:12px;">'
        f'⚠️ 跳过 {stats.skipped} 条</span>'
        if stats.skipped > 0
        else '<span style="display:inline-block;padding:4px 10px;border-radius:999px;'
             'background:#ecfdf5;color:#065f46;font-weight:700;font-size:12px;">✅ 全部导入</span>'
    )
    return f"""\
<!doctype html>
<html><head><meta charset="utf-8"/><title>产品信息库导入</title></head>
<body style="margin:0;padding:0;background:#f6f7fb;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,'PingFang SC','Hiragino Sans GB','Microsoft YaHei',sans-serif;">
  <div style="max-width:920px;margin:0 auto;padding:24px;">
    <div style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;">
      <div style="padding:16px 18px;background:linear-gradient(135deg,#1f2937,#111827);color:#fff;">
        <div style="font-size:16px;font-weight:700;">产品信息库导入通知</div>
        <div style="margin-top:6px;font-size:12px;opacity:.85;">rpa-task · product_sku</div>
      </div>
      <div style="padding:16px 18px;">
        {badge}
        <table style="margin-top:14px;font-size:13px;border-collapse:collapse;">
          <tr><td style="padding:4px 8px;color:#6b7280;">文件</td><td style="padding:4px 8px;color:#111827;font-family:Menlo,Consolas,monospace;">{_esc(stats.file_name)}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">开始</td><td style="padding:4px 8px;color:#111827;">{_esc(stats.started_at)}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">结束</td><td style="padding:4px 8px;color:#111827;">{_esc(stats.ended_at)}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">Excel 数据行</td><td style="padding:4px 8px;color:#111827;">{stats.excel_rows}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">实际插入</td><td style="padding:4px 8px;color:#065f46;font-weight:600;">{stats.inserted}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">已存在且已更新</td><td style="padding:4px 8px;color:#1d4ed8;font-weight:600;">{stats.updated}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">跳过条数</td><td style="padding:4px 8px;color:#b45309;font-weight:600;">{stats.skipped}</td></tr>
        </table>

        {("<h4 style='margin:18px 0 8px 0;color:#111827;'>跳过明细</h4>"
          "<div style='overflow:auto;'>"
          "<table style='width:100%;border-collapse:collapse;font-size:12.5px;'>"
          "<thead>"
          "<tr style='background:#f9fafb;color:#374151;'>"
          "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:right;'>Excel行</th>"
          "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>产品编码 product_sku</th>"
          "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>商品ID product_uid</th>"
          "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>供应商</th>"
          "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>颜色</th>"
          "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>原因</th>"
          "</tr></thead><tbody>"
          + rows_html +
          "</tbody></table></div>") if stats.skipped > 0 else ""}
      </div>
    </div>
    <div style="text-align:center;color:#9ca3af;font-size:11px;margin-top:12px;">Generated by rpa-task / import_product_master.py</div>
  </div>
</body></html>
"""


def _build_skip_text(stats: ImportStats) -> str:
    head = (
        f"产品信息库导入通知\n"
        f"文件：{stats.file_name}\n"
        f"开始：{stats.started_at}    结束：{stats.ended_at}\n"
        f"Excel 数据行：{stats.excel_rows}    插入：{stats.inserted}    更新：{stats.updated}    跳过：{stats.skipped}\n"
    )
    if stats.skipped == 0:
        return head + "\n全部导入成功。\n"
    lines = ["", "跳过明细：", "Excel行  product_sku  product_uid  供应商  颜色  原因"]
    for r in stats.skipped_records:
        lines.append(
            f"{r.excel_row}  {r.product_sku}  {r.product_uid or ''}  "
            f"{r.supplier_name or ''}  {r.product_color or ''}  {r.reason}"
        )
    return head + "\n".join(lines) + "\n"


def _send_notification(stats: ImportStats) -> bool:
    """
    发送邮件。返回是否成功发送。
    缺少配置或无收件人时静默跳过（不抛异常）。
    """
    host = os.getenv("SMTP_HOST", "").strip()
    port_raw = os.getenv("SMTP_PORT", "587").strip()
    user = os.getenv("SMTP_USER") or None
    pwd = os.getenv("SMTP_PASS") or None
    use_starttls = _env_bool("SMTP_STARTTLS", True)
    use_ssl = _env_bool("SMTP_SSL", False)
    mail_from = os.getenv("MAIL_FROM", "").strip() or (user or "")
    mail_to_raw = os.getenv("MAIL_TO", "").strip()
    subject_prefix = os.getenv("MAIL_SUBJECT", "").strip() or "rpa-task"

    if not host or not mail_from or not mail_to_raw:
        _LOG.warn("缺少 SMTP/邮件配置（SMTP_HOST/MAIL_FROM/MAIL_TO），跳过邮件通知")
        return False

    try:
        port = int(port_raw)
    except ValueError:
        port = 587

    mail_to = [s.strip() for s in mail_to_raw.replace(";", ",").split(",") if s.strip()]
    if not mail_to:
        _LOG.warn("MAIL_TO 为空，跳过邮件通知")
        return False

    if stats.skipped > 0:
        subject = f"[{subject_prefix}] 产品信息库导入 - 跳过 product_sku 重复 {stats.skipped} 条"
    else:
        subject = (
            f"[{subject_prefix}] 产品信息库导入 - 全部成功（插入 {stats.inserted}，更新 {stats.updated}）"
        )

    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = ", ".join(mail_to)
    msg["Subject"] = subject
    msg.set_content(_build_skip_text(stats))
    msg.add_alternative(_build_skip_html(stats), subtype="html")

    try:
        if use_ssl:
            with smtplib.SMTP_SSL(host, port, timeout=30) as server:
                if user and pwd:
                    server.login(user, pwd)
                server.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=30) as server:
                server.ehlo()
                if use_starttls:
                    server.starttls()
                    server.ehlo()
                if user and pwd:
                    server.login(user, pwd)
                server.send_message(msg)
        _LOG.info(f"邮件已发送：to={mail_to} subject={subject!r}")
        return True
    except Exception as e:
        _LOG.error(f"邮件发送失败：{type(e).__name__}: {e}")
        return False


# ============================================================
# 主流程
# ============================================================
def default_excel_path() -> Path:
    return v2_root().parent / "excel" / "base" / "产品信息库.xlsx"


def import_file(conn, xlsx: Path) -> ImportStats:
    stats = ImportStats(file_name=xlsx.name, started_at=datetime.now().isoformat(timespec="seconds"))
    _LOG.warn(f"读取 Excel：{xlsx} sheet={SHEET_NAME!r}")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"未找到 sheet「{SHEET_NAME}」，可用：{wb.sheetnames}")
    ws = wb[SHEET_NAME]

    col_index = _build_col_index(ws)
    resolved = _resolve_columns(col_index)
    if "product_sku" not in resolved:
        raise RuntimeError("Excel 表头未找到「产品编码」列，无法继续")

    # ===== 第一遍：扫描所有数据行，做行级转换 + Excel 内部去重 =====
    seen_sku: set[str] = set()
    candidates: list[dict[str, Any]] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        d = _row_to_dict(row, resolved)
        sku = d.get("product_sku")
        # 行内全部业务字段为空 → 视为空行
        if all(v is None for k, v in d.items() if k != "source_type"):
            continue
        stats.excel_rows += 1
        if not sku:
            stats.skipped_records.append(
                SkipRecord(
                    excel_row=ridx,
                    product_sku="",
                    product_uid=d.get("product_uid"),
                    supplier_name=d.get("supplier_name"),
                    product_color=d.get("product_color"),
                    reason="缺产品编码",
                )
            )
            continue
        if sku in seen_sku:
            stats.skipped_records.append(
                SkipRecord(
                    excel_row=ridx,
                    product_sku=sku,
                    product_uid=d.get("product_uid"),
                    supplier_name=d.get("supplier_name"),
                    product_color=d.get("product_color"),
                    reason="Excel内重复",
                )
            )
            continue
        seen_sku.add(sku)
        d["_excel_row"] = ridx
        candidates.append(d)

    _LOG.info(
        f"扫描完成：Excel数据行={stats.excel_rows} 候选行={len(candidates)} "
        f"已跳过={stats.skipped}"
    )

    # ===== 第二遍：批量查询 DB 已存在的 product_sku =====
    existed = _existing_product_skus(conn, [d["product_sku"] for d in candidates])
    _LOG.info(f"DB 内已存在 product_sku：{len(existed)} 个")
    existing_rows = _fetch_rows_for_update(conn, list(existed))

    # ===== 第三遍：新行 INSERT；已存在且允许列有变化则 UPDATE =====
    rows_to_insert: list[tuple[Any, ...]] = []
    rows_to_update: list[tuple[Any, ...]] = []
    for d in candidates:
        sku = d["product_sku"]
        d["line_hash"] = stable_line_hash(row_subset_for_line_hash(d, LINE_HASH_KEYS))
        if sku in existed:
            old = existing_rows.get(sku)
            if old is None:
                rows_to_insert.append(tuple(d[c] for c in INSERT_COLS))
                continue
            if not _needs_update(old, d):
                continue
            vals = tuple(d[c] for c in UPDATABLE_COLS) + (sku,)
            rows_to_update.append(vals)
            continue
        rows_to_insert.append(tuple(d[c] for c in INSERT_COLS))

    _LOG.info(f"准备 INSERT IGNORE：{len(rows_to_insert)} 行，UPDATE：{len(rows_to_update)} 行")
    stats.inserted = _insert_rows(conn, rows_to_insert)
    stats.updated = _update_rows(conn, rows_to_update)
    stats.ended_at = datetime.now().isoformat(timespec="seconds")
    _LOG.info(
        f"导入完成：Excel数据行={stats.excel_rows} 插入={stats.inserted} 更新={stats.updated} 跳过={stats.skipped}"
    )
    return stats


def main() -> int:
    setup_stdout_utf8()
    _load_env_files()

    ap = argparse.ArgumentParser(description="导入 产品信息库.xlsx -> product_sku 表")
    ap.add_argument("--file", type=Path, default=None, help="指定 xlsx，默认 python/excel/base/产品信息库.xlsx")
    ap.add_argument("--no-mail", action="store_true", help="即便有跳过也不发邮件")
    ap.add_argument("--always-mail", action="store_true", help="无跳过也发一封成功邮件")
    ap.add_argument(
        "--quiet-on-db-existed",
        action="store_true",
        help="若所有跳过都是历史原因「DB内已存在」，不发邮件；"
             "当前逻辑下重复 SKU 会更新库内字段，一般仅 Excel 内重复/缺编码 会触发跳过邮件",
    )
    args = ap.parse_args()

    xlsx = args.file or default_excel_path()
    if not xlsx.is_file():
        _LOG.error(f"Excel 不存在：{xlsx}")
        return 2

    cfg = load_db_config()
    _LOG.info(f"连接数据库：host={cfg.host} port={cfg.port} database={cfg.database} user={cfg.user}")
    conn = connect(cfg)
    try:
        stats = import_file(conn, xlsx)
        conn.commit()
        _LOG.info(
            f"已提交事务：插入={stats.inserted} 更新={stats.updated} 跳过={stats.skipped} 跳过原因={stats.skip_reasons}"
        )
    except Exception:
        conn.rollback()
        _LOG.error("发生异常，已回滚事务")
        raise
    finally:
        conn.close()
        _LOG.info("数据库连接已关闭")

    # ===== 邮件通知决策 =====
    # 优先级（从高到低）：
    #   1. --no-mail        → 任何情况都不发
    #   2. --always-mail    → 任何情况都发（含 0 跳过的成功邮件）
    #   3. --quiet-on-db-existed + 所有跳过都是「DB内已存在」 → 不发（重复导入场景）
    #   4. 有跳过           → 发
    #   5. 无跳过           → 不发
    if args.no_mail:
        _LOG.info("--no-mail 已设置，跳过邮件通知")
    elif args.always_mail:
        _send_notification(stats)
    elif args.quiet_on_db_existed and stats.has_only_db_existed_skips:
        _LOG.info(
            f"--quiet-on-db-existed 已设置，且全部跳过均为「DB内已存在」（{stats.skipped} 条），"
            f"按重复导入场景静默处理"
        )
    elif stats.skipped > 0:
        _send_notification(stats)
    else:
        _LOG.info("无跳过记录，按默认策略不发邮件（如需总是发送请加 --always-mail）")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
