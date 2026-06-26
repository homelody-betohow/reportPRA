from __future__ import annotations

"""
TEMU 订单费用：分两步执行（可独立运行）

步骤 1 — 导入 Excel → 表 temu_order_detail（UPSERT，按 line_hash）
  - 仅解析「TEMU-订单详情.xlsx」并落库，不在此步查询 sales_order_shipped。
  - order_no、shop_name_en、shop_alias、platform_site、warehouse_sku、product_sku 一律写空串（不写入 Excel 中的订单号/店铺列）。

步骤 2 — 从 temu_order_detail 补全并发货表费用
  - 2a：对「参考号有效且上述字段仍为空或旧占位 '-'」的明细行，按 ref_no + SKU 多候选匹配
        sales_order_shipped（platform=semitemu），从发货表回写这六个字段。
  - 2b：用 temu_order_detail 补全发货表费用（紫鸟 sheet：`platform_shipping_pay` = 运费回款+税金收入+运费税收入+预估扣除，
        与 Excel 各列一致；明细表里上述四列分别落库，不合并到 shipping_receipt）。

运行示例（在 python/ 目录下）：
  python v2/orders/import_temu_fee.py
  python v2/orders/import_temu_fee.py --file path/to/TEMU-订单详情.xlsx
  python v2/orders/import_temu_fee.py --only-step1    # 只导入明细表
  python v2/orders/import_temu_fee.py --only-step2    # 只执行补全 + 回填 shipped（无需 Excel）
  python v2/orders/import_temu_fee.py --dry-run
  python v2/orders/import_temu_fee.py --no-detail-table   # 等同 --only-step2（兼容旧参数）
  python v2/orders/import_temu_fee.py --write-order-days 30   # 仅写入近 30 天（见参数说明）

表结构：docs/database/030_temu_order_detail.sql

邮件：未匹配明细表仅列出「发货时间 / 明细订单时间」在最近 MAIL_NOTIFY_ORDER_DAYS（默认 60）天内的行；
      主题与摘要中的「近 N 天」计数与此一致，全量条数见正文说明。
"""

import argparse
import os
import re
import smtplib
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from email.message import EmailMessage
from pathlib import Path
from typing import Any

_ORDERS_DIR = Path(__file__).resolve().parent
_V2_DIR = _ORDERS_DIR.parent
_WR_DIR = _V2_DIR / "warehouse-rent"
for _p in (_ORDERS_DIR, _V2_DIR, _WR_DIR):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

import openpyxl  # type: ignore[import-untyped]

from config.fx_rates import FxRates, SYMBOL_TO_ISO, format_summary, load_rates
from db import connect, load_db_config
from excel_common import cell_dt, default_order_excel_dir, row_subset_for_line_hash, stable_line_hash, upsert_rows
from logger import get_logger, setup_stdout_utf8

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None  # type: ignore[assignment]


TABLE = "sales_order_shipped"
TEMU_DETAIL_TABLE = "temu_order_detail"
PLATFORM_FILTER = "semitemu"
BASE_CURRENCY = "EUR"
DEFAULT_FILE_NAME = "TEMU-订单详情.xlsx"
# 邮件中「未匹配」明细表仅展示该时间窗内的订单（发货时间 / 明细订单时间）
MAIL_NOTIFY_ORDER_DAYS = 60
_LOG = get_logger("TEMU-FEE")

SHEET_TO_CURRENCY: dict[str, str] = {
    "AIHOMEU": "RMB",
    "BathVogue_EU": "RMB",
    "HAUSE_MATE": "USD",
    "KR-A": "USD",
    "KR-B": "USD",
    "KR-C": "USD",
    "HJ-A": "USD",
    "HJ-B": "USD",
    "HJ-C": "USD",
    "NF-A": "USD",
    "NF-B": "USD",
    "NF-C": "USD",
    "TEMU-AL": "ZINIAO",
    "TEMU-BZ": "ZINIAO",
    "TEMU-AQ": "ZINIAO",
}

_TEMU_WH_PREFIX = "900008-"
_NUM_PAT = re.compile(r"(\d+(?:[.,]\d+)?)")


@dataclass
class UnmatchedDb:
    id: int
    order_no: str | None
    ref_no: str | None
    platform_sku_orig: str | None
    warehouse_sku: str | None
    warehouse_sku_qty: int | None
    ship_time: datetime | None = None  # sales_order_shipped.ship_time，用于邮件按天筛选


@dataclass
class UnmatchedExcel:
    """步骤 2 中：temu_order_detail 有价格但未命中任何 shipped 行的记录（展示用）。"""
    sheet: str
    excel_row: int
    ref_no: str
    sku_key: str
    pay_currency: str
    unit_price_pay: Decimal
    unit_price_base: Decimal
    order_time: datetime | None = None  # temu_order_detail.order_time，用于邮件按天筛选


@dataclass
class ImportStats:
    file_name: str = ""
    started_at: str = ""
    ended_at: str = ""
    steps_label: str = ""  # 如 "1+2" / "1" / "2"
    excel_total_rows: int = 0
    excel_cancelled: int = 0
    excel_invalid: int = 0
    excel_valid_rows: int = 0
    excel_unique_keys: int = 0
    db_temu_rows: int = 0
    matched: int = 0
    updated: int = 0
    detail_rows_saved: int = 0
    detail_backfilled: int = 0  # 步骤 2a：从 shipped 写回 temu_order_detail 的行数
    write_order_days_applied: int | None = None  # 若启用「近 N 天」写入则记录 N
    excel_rows_skipped_write_window: int = 0  # 步骤 1：因 order_time 落在窗口外而未写入的行数
    unmatched_db: list[UnmatchedDb] = field(default_factory=list)
    unmatched_excel: list[UnmatchedExcel] = field(default_factory=list)


def _coerce_datetime_for_notify(v: Any) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return None


def _mail_notify_cutoff() -> datetime:
    d = date.today() - timedelta(days=MAIL_NOTIFY_ORDER_DAYS)
    return datetime(d.year, d.month, d.day)


def _in_mail_notify_window(dt: Any, *, cutoff: datetime) -> bool:
    ts = _coerce_datetime_for_notify(dt)
    if ts is None:
        return False
    return ts >= cutoff


def _unmatched_shipped_row(r: dict[str, Any], *, ref_no: str | None = None) -> UnmatchedDb:
    try:
        rid = int(r["id"])
    except (TypeError, ValueError, KeyError):
        rid = 0
    return UnmatchedDb(
        id=rid,
        order_no=r.get("order_no"),
        ref_no=ref_no if ref_no is not None else r.get("ref_no"),
        platform_sku_orig=r.get("platform_sku_orig"),
        warehouse_sku=r.get("warehouse_sku"),
        warehouse_sku_qty=r.get("warehouse_sku_qty"),
        ship_time=_coerce_datetime_for_notify(r.get("ship_time")),
    )


def _write_window_start(*, days: int) -> datetime:
    """今天 0 点往前推 days 天（含当天）：order_time / ship_time >= 该时刻则视为在窗口内。"""
    d = date.today() - timedelta(days=days)
    return datetime(d.year, d.month, d.day)


def _naive_detail_order_time(ot: Any) -> datetime | None:
    if isinstance(ot, datetime):
        return ot.replace(tzinfo=None) if ot.tzinfo else ot
    if isinstance(ot, date):
        return datetime(ot.year, ot.month, ot.day)
    return None


def _is_placeholder_detail_order_time(ot: Any) -> bool:
    """缺少日期列时脚本使用 2000-01-01，不参与「近 N 天」写入窗口。"""
    t = _naive_detail_order_time(ot)
    if t is None:
        return True
    return t.date() <= date(2000, 1, 1)


def _detail_row_in_write_window(d: dict[str, Any], *, window_start: datetime) -> bool:
    if _is_placeholder_detail_order_time(d.get("order_time")):
        return False
    t = _naive_detail_order_time(d.get("order_time"))
    return t is not None and t >= window_start


def _unmatched_lists_for_mail(
    stats: ImportStats,
) -> tuple[list[UnmatchedDb], list[UnmatchedExcel], datetime]:
    """邮件列表：仅含发货时间/订单时间在最近 MAIL_NOTIFY_ORDER_DAYS 天内的未匹配行。"""
    cutoff = _mail_notify_cutoff()
    db_rows = [x for x in stats.unmatched_db if _in_mail_notify_window(x.ship_time, cutoff=cutoff)]
    ex_rows = [x for x in stats.unmatched_excel if _in_mail_notify_window(x.order_time, cutoff=cutoff)]
    return db_rows, ex_rows, cutoff


# ---------------------------------------------------------------------------
# 通用解析
# ---------------------------------------------------------------------------


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _to_decimal_simple(v: Any) -> Decimal | None:
    if _is_blank(v):
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = str(v).strip().replace(",", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_ziniao_amount(v: Any, fx: FxRates) -> Decimal | None:
    if _is_blank(v):
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = str(v).replace(" ", "").replace("\u00a0", "")
    iso_ccy = "EUR"
    for sym, iso in SYMBOL_TO_ISO:
        if sym in s:
            iso_ccy = iso
            break
    m = _NUM_PAT.search(s)
    if not m:
        return None
    num_str = m.group(1).replace(",", ".")
    try:
        num = Decimal(num_str)
    except InvalidOperation:
        return None
    return fx.to_eur(num, iso_ccy)


def _ref_no(v: Any) -> str | None:
    if _is_blank(v):
        return None
    return str(v).strip()


def _sku_key(v: Any) -> str | None:
    if _is_blank(v):
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _product_sku_from_wh(warehouse_sku: str) -> str:
    s = (warehouse_sku or "").strip()
    if not s:
        return ""
    if s.startswith(_TEMU_WH_PREFIX):
        tail = s[len(_TEMU_WH_PREFIX) :].strip()
        return (tail if tail else s)[:128]
    return s[:128]


def _cell_dt_excel(v: Any) -> datetime | None:
    if _is_blank(v):
        return None
    if isinstance(v, datetime):
        return v.replace(microsecond=0) if v.microsecond else v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            from openpyxl.utils.datetime import from_excel

            dt_any = from_excel(float(v))
            if isinstance(dt_any, datetime):
                return dt_any.replace(microsecond=0) if dt_any.microsecond else dt_any
            if isinstance(dt_any, date):
                return datetime(dt_any.year, dt_any.month, dt_any.day)
        except Exception:
            pass
    return cell_dt(v)


def _decimal_cell(row_tuple: tuple[Any, ...], idx: int | None) -> Decimal | None:
    if idx is None or idx >= len(row_tuple):
        return None
    return _to_decimal_simple(row_tuple[idx])


def _fx_rate_to_base(fx: FxRates, pay_currency: str | None) -> Decimal | None:
    if not pay_currency or not str(pay_currency).strip():
        return None
    c = str(pay_currency).strip().upper()
    if c == "EUR":
        return Decimal("1")
    if c == "CNY":
        return (Decimal("1") / fx.rmb_per_eur).quantize(Decimal("0.00000001"))
    if c in fx.rates_to_eur:
        return fx.rates_to_eur[c]
    return None


def _detail_money_to_decimal(v: Any) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return Decimal("0")


def _platform_shipping_pay_from_detail(hit: dict[str, Any]) -> Decimal:
    """
    回填 sales_order_shipped.platform_shipping_pay：
    - ZINIAO：与 v1 一致，为 运费回款 + 税金收入 + 运费税收入 + 预估扣除（多列在库内分别存）；
    - RMB/USD：仅 shipping_receipt（即 Excel 运费回款）。
    """
    ship = _detail_money_to_decimal(hit.get("shipping_receipt"))
    if str(hit.get("currency_group") or "").strip().upper() == "ZINIAO":
        return (
            ship
            + _detail_money_to_decimal(hit.get("tax_income"))
            + _detail_money_to_decimal(hit.get("shipping_tax_income"))
            + _detail_money_to_decimal(hit.get("deduction_estimate"))
        ).quantize(Decimal("0.000001"))
    return ship.quantize(Decimal("0.000001"))


# ---------------------------------------------------------------------------
# 步骤 1：Excel → temu_order_detail
# ---------------------------------------------------------------------------

DETAIL_LINE_HASH_KEYS: tuple[str, ...] = (
    "source_file",
    "excel_sheet",
    "excel_row",
    "ref_no",
    "sku_key",
)

TEMU_DETAIL_INSERT_COLS: tuple[str, ...] = (
    "line_hash",
    "excel_sheet",
    "currency_group",
    "pay_currency",
    "order_time",
    "order_no",
    "ref_no",
    "sku_key",
    "warehouse_sku",
    "product_sku",
    "sku_quantity",
    "unit_price_pay",
    "sales_receipt",
    "sales_reverse",
    "shipping_receipt",
    "tax_income",
    "shipping_tax_income",
    "deduction_estimate",
    "income_estimate",
    "shop_name_en",
    "shop_alias",
    "platform_site",
    "row_kind",
)


def _read_sheet_detail_rows(
    ws: Any,
    sheet_currency: str,
    fx: FxRates,
    stats: ImportStats,
    source_file: str,
) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [(h.strip() if isinstance(h, str) else h) for h in rows[0]]
    name = ws.title

    def col(name_: str) -> int | None:
        try:
            return headers.index(name_)
        except ValueError:
            return None

    def first_col(*candidates: str) -> int | None:
        for cn in candidates:
            try:
                return headers.index(cn)
            except ValueError:
                continue
        return None

    i_ref = col("参考号")
    i_price = col("产品单价")
    i_ship = col("运费回款")
    i_order_time = first_col("订单时间", "日期")
    i_qty = col("购买数量") or col("数量")
    i_sales = first_col("销售回款", "销售收入")
    i_sales_rev = first_col("销售冲回", "销售冲红")
    i_tax = col("税金收入")
    i_ship_tax = col("运费税收入")
    i_deduct = col("预估扣除金额")
    i_income_est = col("预计收入")

    is_ziniao = sheet_currency == "ZINIAO"
    if is_ziniao:
        i_sku = col("SKU 货号")
        i_tax_inc = col("税金收入")
        i_freight_tax = col("运费税收入")
        i_estimate_deduct = col("预估扣除金额")
    else:
        i_sku = col("SKU ID")
        i_tax_inc = i_freight_tax = i_estimate_deduct = None

    if i_ref is None or i_sku is None or i_price is None:
        _LOG.warn(f"sheet「{name}」缺少必要列（参考号/SKU/产品单价），跳过")
        return []

    if i_order_time is None:
        _LOG.warn(f"sheet「{name}」未找到日期列（订单时间/日期），order_time 将回退为 2000-01-01")

    out: list[dict[str, Any]] = []

    for ridx, row in enumerate(rows[1:], start=2):
        row_tuple = tuple(row) if row is not None else tuple()
        ref = _ref_no(row_tuple[i_ref] if i_ref < len(row_tuple) else None)
        sku = _sku_key(row_tuple[i_sku] if i_sku < len(row_tuple) else None)
        price_raw = row_tuple[i_price] if i_price < len(row_tuple) else None
        ship_raw = row_tuple[i_ship] if (i_ship is not None and i_ship < len(row_tuple)) else None

        if ref is None and sku is None and _is_blank(price_raw):
            continue

        # 非「客户取消订单」：购买数量列为空或解析为 0 及以下的行整行跳过（不计入 excel_total_rows）
        is_cancelled = isinstance(price_raw, str) and price_raw.strip() == "客户取消订单"
        if not is_cancelled and i_qty is not None:
            raw_qty = row_tuple[i_qty] if i_qty < len(row_tuple) else None
            if _is_blank(raw_qty):
                continue
            qd = _decimal_cell(row_tuple, i_qty)
            if qd is None:
                continue
            try:
                qn = int(qd)
            except (TypeError, ValueError, InvalidOperation):
                continue
            if qn <= 0:
                continue

        stats.excel_total_rows += 1

        ot = _cell_dt_excel(
            row_tuple[i_order_time] if i_order_time is not None and i_order_time < len(row_tuple) else None
        )
        order_time = ot or datetime(2000, 1, 1)
        qty_dec = _decimal_cell(row_tuple, i_qty)
        try:
            sku_qty = int(qty_dec) if qty_dec is not None else 1
        except (TypeError, ValueError, InvalidOperation):
            sku_qty = 1
        sku_qty = max(0, min(sku_qty, 65535))

        base_common: dict[str, Any] = {
            "source_file": source_file,
            "excel_sheet": name[:128],
            "currency_group": sheet_currency[:16],
            "order_time": order_time,
            "order_no": "",
            "ref_no": ref,
            "sku_key": sku,
            "warehouse_sku": "",
            "product_sku": "",
            "sku_quantity": sku_qty,
            "shop_name_en": "",
            "shop_alias": "",
            "platform_site": "",
        }

        if is_ziniao:
            sales_recv_row = _parse_ziniao_amount(
                row_tuple[i_sales] if i_sales is not None and i_sales < len(row_tuple) else None,
                fx,
            )
            sales_rev_row = _parse_ziniao_amount(
                row_tuple[i_sales_rev] if i_sales_rev is not None and i_sales_rev < len(row_tuple) else None,
                fx,
            )
        else:
            sales_recv_row = _decimal_cell(row_tuple, i_sales)
            sales_rev_row = _decimal_cell(row_tuple, i_sales_rev)

        if isinstance(price_raw, str) and price_raw.strip() == "客户取消订单":
            stats.excel_cancelled += 1
            base_common["sku_quantity"] = 0
            out.append(
                {
                    **base_common,
                    "excel_row": ridx,
                    "row_kind": "cancelled",
                    "pay_currency": None,
                    "unit_price_pay": None,
                    "sales_receipt": sales_recv_row,
                    "sales_reverse": sales_rev_row,
                    "shipping_receipt": None,
                    "tax_income": None,
                    "shipping_tax_income": None,
                    "deduction_estimate": None,
                    "income_estimate": None,
                }
            )
            continue

        if ref is None or sku is None:
            stats.excel_invalid += 1
            out.append(
                {
                    **base_common,
                    "excel_row": ridx,
                    "row_kind": "invalid_missing_key",
                    "pay_currency": None,
                    "unit_price_pay": None,
                    "sales_receipt": sales_recv_row,
                    "sales_reverse": sales_rev_row,
                    "shipping_receipt": None,
                    "tax_income": None,
                    "shipping_tax_income": None,
                    "deduction_estimate": None,
                    "income_estimate": None,
                }
            )
            continue

        if is_ziniao:
            pay_currency = "EUR"
            unit_price_pay_dec = _parse_ziniao_amount(price_raw, fx)
            if unit_price_pay_dec is None:
                stats.excel_invalid += 1
                out.append(
                    {
                        **base_common,
                        "excel_row": ridx,
                        "row_kind": "invalid_parse",
                        "pay_currency": None,
                        "unit_price_pay": None,
                        "sales_receipt": sales_recv_row,
                        "sales_reverse": sales_rev_row,
                        "shipping_receipt": None,
                        "tax_income": None,
                        "shipping_tax_income": None,
                        "deduction_estimate": None,
                        "income_estimate": None,
                    }
                )
                continue
            ship_base = _parse_ziniao_amount(ship_raw, fx) or Decimal("0")
            tax_inc = (
                _parse_ziniao_amount(
                    row_tuple[i_tax_inc] if (i_tax_inc is not None and i_tax_inc < len(row_tuple)) else None,
                    fx,
                )
                or Decimal("0")
            )
            freight_tax = (
                _parse_ziniao_amount(
                    row_tuple[i_freight_tax]
                    if (i_freight_tax is not None and i_freight_tax < len(row_tuple))
                    else None,
                    fx,
                )
                or Decimal("0")
            )
            est_deduct = (
                _parse_ziniao_amount(
                    row_tuple[i_estimate_deduct]
                    if (i_estimate_deduct is not None and i_estimate_deduct < len(row_tuple))
                    else None,
                    fx,
                )
                or Decimal("0")
            )
            shipping_pay_dec = (ship_base + tax_inc + freight_tax + est_deduct).quantize(Decimal("0.000001"))
        else:
            unit_price_pay_dec = _to_decimal_simple(price_raw)
            if unit_price_pay_dec is None:
                stats.excel_invalid += 1
                out.append(
                    {
                        **base_common,
                        "excel_row": ridx,
                        "row_kind": "invalid_parse",
                        "pay_currency": None,
                        "unit_price_pay": None,
                        "sales_receipt": sales_recv_row,
                        "sales_reverse": sales_rev_row,
                        "shipping_receipt": None,
                        "tax_income": None,
                        "shipping_tax_income": None,
                        "deduction_estimate": None,
                        "income_estimate": None,
                    }
                )
                continue
            shipping_pay_dec = _to_decimal_simple(ship_raw) or Decimal("0")
            if sheet_currency == "RMB":
                pay_currency = "CNY"
            elif sheet_currency == "USD":
                pay_currency = "USD"
            else:
                _LOG.warn(f"sheet「{name}」未知 sheet_currency={sheet_currency}，跳过")
                stats.excel_invalid += 1
                out.append(
                    {
                        **base_common,
                        "excel_row": ridx,
                        "row_kind": "invalid_parse",
                        "pay_currency": None,
                        "unit_price_pay": None,
                        "sales_receipt": sales_recv_row,
                        "sales_reverse": sales_rev_row,
                        "shipping_receipt": None,
                        "tax_income": None,
                        "shipping_tax_income": None,
                        "deduction_estimate": None,
                        "income_estimate": None,
                    }
                )
                continue

        stats.excel_valid_rows += 1
        if is_ziniao:
            # 明细表与 Excel 列一一对应；勿把「四者之和」写入 shipping_receipt
            sh_rec = ship_base
            tx = tax_inc
            stx = freight_tax
            ded = est_deduct
        else:
            sh_rec = shipping_pay_dec
            tx = _decimal_cell(row_tuple, i_tax)
            stx = _decimal_cell(row_tuple, i_ship_tax)
            ded = _decimal_cell(row_tuple, i_deduct)
        inc_est = _decimal_cell(row_tuple, i_income_est)
        out.append(
            {
                **base_common,
                "excel_row": ridx,
                "row_kind": "valid",
                "pay_currency": pay_currency,
                "unit_price_pay": unit_price_pay_dec,
                "sales_receipt": sales_recv_row,
                "sales_reverse": sales_rev_row,
                "shipping_receipt": sh_rec,
                "tax_income": tx,
                "shipping_tax_income": stx,
                "deduction_estimate": ded,
                "income_estimate": inc_est,
            }
        )
    return out


def _read_temu_excel(xlsx: Path, fx: FxRates, stats: ImportStats) -> list[dict[str, Any]]:
    _LOG.warn(f"读取 Excel：{xlsx}")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    all_detail: list[dict[str, Any]] = []
    src = xlsx.name
    for name in wb.sheetnames:
        sheet_currency = SHEET_TO_CURRENCY.get(name)
        if sheet_currency is None:
            _LOG.warn(f"sheet「{name}」未在 SHEET_TO_CURRENCY 已知分组，跳过")
            continue
        ws = wb[name]
        chunk = _read_sheet_detail_rows(ws, sheet_currency, fx, stats, src)
        all_detail.extend(chunk)
        _LOG.info(f"sheet「{name}」({sheet_currency}) 明细行={len(chunk)}")

    uniq: set[tuple[str, str]] = set()
    dup = 0
    for d in all_detail:
        if d.get("row_kind") != "valid":
            continue
        ref = str(d.get("ref_no") or "").strip()
        sk = str(d.get("sku_key") or "").strip()
        if not ref or not sk:
            continue
        k = (ref, sk)
        if k in uniq:
            dup += 1
        uniq.add(k)
    stats.excel_unique_keys = len(uniq)
    if dup:
        _LOG.warn(f"Excel 内 (参考号, SKU) 有重复 {dup} 条，统计唯一键时已去重")
    return all_detail


def _save_temu_order_detail(conn: Any, detail_rows: list[dict[str, Any]]) -> int:
    if not detail_rows:
        return 0
    tuples: list[tuple[Any, ...]] = []
    for d in detail_rows:
        h_in = row_subset_for_line_hash(d, DETAIL_LINE_HASH_KEYS)
        d_copy = dict(h_in)
        if d_copy.get("ref_no") is None:
            d_copy["ref_no"] = ""
        if d_copy.get("sku_key") is None:
            d_copy["sku_key"] = ""
        lh = stable_line_hash(d_copy)
        tuples.append(
            (
                lh,
                d["excel_sheet"],
                d["currency_group"],
                d.get("pay_currency"),
                d["order_time"],
                d["order_no"],
                d.get("ref_no"),
                d.get("sku_key"),
                d["warehouse_sku"],
                d["product_sku"],
                int(d["sku_quantity"]),
                d.get("unit_price_pay"),
                d.get("sales_receipt"),
                d.get("sales_reverse"),
                d.get("shipping_receipt"),
                d.get("tax_income"),
                d.get("shipping_tax_income"),
                d.get("deduction_estimate"),
                d.get("income_estimate"),
                d["shop_name_en"],
                d["shop_alias"],
                d["platform_site"],
                d["row_kind"],
            )
        )
    _LOG.info(f"准备 UPSERT `{TEMU_DETAIL_TABLE}`：{len(tuples)} 行")
    return upsert_rows(conn, table=TEMU_DETAIL_TABLE, columns=list(TEMU_DETAIL_INSERT_COLS), rows=tuples)


def step1_import_excel_to_detail(
    conn: Any,
    xlsx: Path,
    fx: FxRates,
    stats: ImportStats,
    *,
    dry_run: bool,
    omit_unit_price_pay: bool = False,
    write_window_start: datetime | None = None,
) -> None:
    detail_rows = _read_temu_excel(xlsx, fx, stats)
    if write_window_start is not None:
        before = len(detail_rows)
        detail_rows = [d for d in detail_rows if _detail_row_in_write_window(d, window_start=write_window_start)]
        stats.excel_rows_skipped_write_window += before - len(detail_rows)
        _LOG.info(
            f"按 order_time≥{write_window_start.date()} 过滤明细写入：保留 {len(detail_rows)}/{before} 行"
            f"（跳过窗口外或缺少有效订单日期 {stats.excel_rows_skipped_write_window} 行）"
        )
    if omit_unit_price_pay:
        for d in detail_rows:
            d["unit_price_pay"] = None
    _LOG.info(
        f"Excel 汇总：总行={stats.excel_total_rows} 取消={stats.excel_cancelled} "
        f"无效={stats.excel_invalid} 有效={stats.excel_valid_rows} 唯一键={stats.excel_unique_keys}"
    )
    if dry_run:
        _LOG.warn("--dry-run：跳过写入 temu_order_detail")
        return
    try:
        stats.detail_rows_saved = _save_temu_order_detail(conn, detail_rows)
        _LOG.info(f"已 UPSERT `{TEMU_DETAIL_TABLE}`：累计处理行数={stats.detail_rows_saved}（按 line_hash）")
    except Exception as e:
        _LOG.error(
            f"写入 `{TEMU_DETAIL_TABLE}` 失败：{type(e).__name__}: {e}。"
            f"请确认已执行 docs/database/030_temu_order_detail.sql。"
        )
        raise


# ---------------------------------------------------------------------------
# 步骤 2：shipped ↔ temu_order_detail
# ---------------------------------------------------------------------------

_TEMU_SHIPPED_FETCH: tuple[str, ...] = (
    "id",
    "ship_time",
    "order_no",
    "ref_no",
    "warehouse_sku",
    "shop_name_en",
    "shop_alias",
    "platform_site",
    "platform_sku_orig",
    "platform_sku",
)


def _fetch_shipped_by_refs_temu(
    conn: Any,
    ref_nos: frozenset[str],
    *,
    chunk_size: int = 200,
) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = defaultdict(list)
    if not ref_nos:
        return out
    cols = ", ".join(f"`{c}`" for c in _TEMU_SHIPPED_FETCH)
    cur = conn.cursor(dictionary=True)
    seen: dict[str, set[Any]] = {}
    try:
        wanted = sorted(ref_nos)
        for i in range(0, len(wanted), chunk_size):
            chunk = wanted[i : i + chunk_size]
            ph = ", ".join(["%s"] * len(chunk))
            sql = f"SELECT {cols} FROM `{TABLE}` WHERE platform = %s AND `ref_no` IN ({ph})"
            cur.execute(sql, (PLATFORM_FILTER, *chunk))
            for raw in cur.fetchall() or []:
                row = {k: raw.get(k) for k in _TEMU_SHIPPED_FETCH}
                rref = str(row.get("ref_no") or "").strip()
                if not rref:
                    continue
                rid = row.get("id")
                bk = seen.setdefault(rref, set())
                dk = rid if rid is not None else id(row)
                if dk in bk:
                    continue
                bk.add(dk)
                out[rref].append(row)
    finally:
        cur.close()
    return out


def _sku_candidates_for_shipped_match(r: dict[str, Any]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def push(s: Any) -> None:
        if s is None:
            return
        t = str(s).strip()
        if not t or t in seen:
            return
        seen.add(t)
        out.append(t)

    push(r.get("platform_sku_orig"))
    push(r.get("platform_sku"))
    push(r.get("warehouse_sku"))
    wh = str(r.get("warehouse_sku") or "").strip()
    if wh.startswith(_TEMU_WH_PREFIX):
        push(wh[len(_TEMU_WH_PREFIX) :].strip())
    return out


def _pick_shipped_for_temu(candidates: list[dict[str, Any]], sku_key: str | None) -> dict[str, Any] | None:
    if not candidates:
        return None
    dedup: dict[Any, dict[str, Any]] = {}
    for r in candidates:
        rid = r.get("id")
        dedup[rid if rid is not None else id(r)] = r
    uniq = list(dedup.values())
    sk = (sku_key or "").strip()
    if sk:
        for r in uniq:
            for cand in _sku_candidates_for_shipped_match(r):
                if cand == sk:
                    return r

    def _ts(r: dict[str, Any]) -> float:
        st = r.get("ship_time")
        if isinstance(st, datetime):
            return st.timestamp()
        return 0.0

    def _rid(r: dict[str, Any]) -> int:
        try:
            return int(r.get("id") or 0)
        except (TypeError, ValueError):
            return 0

    return max(uniq, key=lambda r: (_ts(r), _rid(r)))


def _fetch_detail_rows_need_shipped_backfill(
    conn: Any, *, min_order_time: datetime | None = None
) -> list[dict[str, Any]]:
    sql = f"""
        SELECT id, ref_no, sku_key, order_no, warehouse_sku
          FROM `{TEMU_DETAIL_TABLE}`
         WHERE row_kind = 'valid'
           AND ref_no IS NOT NULL AND TRIM(ref_no) <> ''
           AND (
                TRIM(IFNULL(order_no,'')) IN ('', '-')
             OR TRIM(IFNULL(warehouse_sku,'')) IN ('', '-')
           )
    """
    params: list[Any] = []
    if min_order_time is not None:
        sql += " AND order_time IS NOT NULL AND order_time >= %s"
        params.append(min_order_time)
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(sql, tuple(params))
        return list(cur.fetchall() or [])
    finally:
        cur.close()


def _update_temu_detail_from_shipped(
    conn: Any,
    detail_id: int,
    ship: dict[str, Any],
    *,
    dry_run: bool,
) -> bool:
    ono = str(ship.get("order_no") or "").strip()[:128]
    wh = str(ship.get("warehouse_sku") or "").strip()[:128]
    ps = _product_sku_from_wh(wh)[:128]
    sne = str(ship.get("shop_name_en") or "").strip()[:128]
    sal = str(ship.get("shop_alias") or "").strip()[:128]
    site = str(ship.get("platform_site") or "").strip()[:64]
    sql = (
        f"UPDATE `{TEMU_DETAIL_TABLE}` SET "
        f"`order_no`=%s, `warehouse_sku`=%s, `product_sku`=%s, "
        f"`shop_name_en`=%s, `shop_alias`=%s, `platform_site`=%s "
        f"WHERE `id`=%s"
    )
    if dry_run:
        return True
    cur = conn.cursor()
    try:
        cur.execute(sql, (ono, wh, ps, sne, sal, site, detail_id))
        return cur.rowcount > 0
    finally:
        cur.close()


def step2a_backfill_detail_from_shipped(
    conn: Any, stats: ImportStats, *, dry_run: bool, write_window_start: datetime | None = None
) -> None:
    rows = _fetch_detail_rows_need_shipped_backfill(conn, min_order_time=write_window_start)
    if not rows:
        _LOG.info("步骤 2a：无需要从 shipped 回写 temu_order_detail 的明细行")
        return
    refs = frozenset(str(r["ref_no"]).strip() for r in rows if r.get("ref_no"))
    smap = _fetch_shipped_by_refs_temu(conn, refs)
    n = 0
    for r in rows:
        ref = str(r.get("ref_no") or "").strip()
        sku_key = r.get("sku_key")
        did = int(r["id"])
        ship = _pick_shipped_for_temu(smap.get(ref) or [], sku_key)
        if not ship:
            continue
        if _update_temu_detail_from_shipped(conn, did, ship, dry_run=dry_run):
            n += 1
    stats.detail_backfilled = n
    _LOG.info(f"步骤 2a：从 shipped 回写 temu_order_detail 行数={n}（dry_run={dry_run}）")


def _fetch_db_temu_orders(
    conn: Any,
    *,
    filter_ref_no: str | None = None,
    min_ship_time: datetime | None = None,
) -> list[dict[str, Any]]:
    cur = conn.cursor(dictionary=True)
    cols = "id, order_no, ref_no, platform_sku_orig, platform_sku, warehouse_sku, warehouse_sku_qty, ship_time"
    params: list[Any] = []
    if filter_ref_no:
        sql = f"SELECT {cols} FROM `{TABLE}` WHERE platform = %s AND ref_no = %s"
        params = [PLATFORM_FILTER, filter_ref_no]
    else:
        sql = f"SELECT {cols} FROM `{TABLE}` WHERE platform = %s"
        params = [PLATFORM_FILTER]
    if min_ship_time is not None:
        sql += " AND ship_time IS NOT NULL AND ship_time >= %s"
        params.append(min_ship_time)
    try:
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()
    finally:
        cur.close()
    return rows


def _fetch_temu_price_lookup(
    conn: Any,
) -> tuple[dict[tuple[str, str], dict[str, Any]], list[dict[str, Any]]]:
    """返回 ((ref_no, sku_token) -> 行)、按 id 去重后的明细列表（用于未命中 shipped 统计）。"""
    sql = f"""
        SELECT id, excel_sheet, ref_no, sku_key, product_sku, order_time,
               pay_currency, unit_price_pay, shipping_receipt,
               currency_group, tax_income, shipping_tax_income, deduction_estimate
          FROM `{TEMU_DETAIL_TABLE}`
         WHERE row_kind = 'valid'
           AND ref_no IS NOT NULL AND TRIM(ref_no) <> ''
           AND unit_price_pay IS NOT NULL
         ORDER BY id ASC
    """
    cur = conn.cursor(dictionary=True)
    out: dict[tuple[str, str], dict[str, Any]] = {}
    by_id: dict[int, dict[str, Any]] = {}
    try:
        cur.execute(sql)
        for raw in cur.fetchall() or []:
            rid = raw.get("id")
            try:
                iid = int(rid) if rid is not None else 0
            except (TypeError, ValueError):
                iid = 0
            if iid:
                by_id[iid] = raw
            ref = str(raw.get("ref_no") or "").strip()
            if not ref:
                continue
            sk = str(raw.get("sku_key") or "").strip()
            ps = str(raw.get("product_sku") or "").strip()
            if sk:
                out[(ref, sk)] = raw
            if ps and ps != sk:
                out[(ref, ps)] = raw
    finally:
        cur.close()
    return out, list(by_id.values())


def _detail_hit_for_shipped_row(
    ref_no: str,
    r: dict[str, Any],
    detail_map: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, Any] | None:
    ref = ref_no.strip()
    if not ref:
        return None
    for sku_try in _sku_candidates_for_shipped_match(r):
        hit = detail_map.get((ref, sku_try))
        if hit is not None:
            return hit
    return None


def _update_shipped_rows(
    conn: Any,
    updates: list[
        tuple[
            str,
            Decimal,
            Decimal,
            Decimal,
            Decimal,
            Decimal,
            Decimal,
            Decimal,
            int,
        ]
    ],
) -> int:
    if not updates:
        return 0
    sql = (
        f"UPDATE `{TABLE}` SET "
        f"  `pay_currency`=%s, "
        f"  `unit_price_pay`=%s, "
        f"  `platform_shipping_pay`=%s, "
        f"  `order_total_pay`=%s, "
        f"  `fx_rate_to_base`=%s, "
        f"  `base_currency`='EUR', "
        f"  `unit_price_base`=%s, "
        f"  `platform_shipping_base`=%s, "
        f"  `order_total_base`=%s "
        f"WHERE `id`=%s"
    )
    cur = conn.cursor()
    affected = 0
    chunk = 300
    for i in range(0, len(updates), chunk):
        batch = updates[i : i + chunk]
        cur.executemany(sql, batch)
        affected += cur.rowcount
    cur.close()
    return max(affected, 0)


def step2b_sync_shipped_from_detail(
    conn: Any,
    fx: FxRates,
    stats: ImportStats,
    *,
    dry_run: bool,
    filter_ref_no: str | None,
    write_window_start: datetime | None = None,
) -> None:
    detail_map, detail_unique_rows = _fetch_temu_price_lookup(conn)
    db_rows = _fetch_db_temu_orders(
        conn, filter_ref_no=filter_ref_no, min_ship_time=write_window_start
    )
    stats.db_temu_rows = len(db_rows)
    if filter_ref_no:
        _LOG.info(
            f"步骤 2b：DB platform={PLATFORM_FILTER!r} AND ref_no={filter_ref_no!r} 行数={stats.db_temu_rows}"
        )
    else:
        _LOG.info(f"步骤 2b：DB platform={PLATFORM_FILTER!r} 行数={stats.db_temu_rows}")
    if write_window_start is not None:
        _LOG.info(f"步骤 2b：仅 ship_time≥{write_window_start.date()} 的发货行参与 UPDATE")

    used_detail_ids: set[int] = set()
    updates: list[
        tuple[str, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal, Decimal, int]
    ] = []

    for r in db_rows:
        ref_no = (r.get("ref_no") or "").strip()
        if not ref_no:
            stats.unmatched_db.append(_unmatched_shipped_row(r))
            continue
        hit = _detail_hit_for_shipped_row(ref_no, r, detail_map)
        if hit is None:
            stats.unmatched_db.append(_unmatched_shipped_row(r, ref_no=ref_no))
            continue
        pay_ccy = hit.get("pay_currency")
        if pay_ccy is None or not str(pay_ccy).strip():
            stats.unmatched_db.append(_unmatched_shipped_row(r, ref_no=ref_no))
            continue
        pay_currency = str(pay_ccy).strip()
        unit_price_pay = hit["unit_price_pay"]
        if not isinstance(unit_price_pay, Decimal):
            try:
                unit_price_pay = Decimal(str(unit_price_pay))
            except InvalidOperation:
                stats.unmatched_db.append(_unmatched_shipped_row(r, ref_no=ref_no))
                continue
        shipping_pay = _platform_shipping_pay_from_detail(hit)
        fx_rate = _fx_rate_to_base(fx, pay_currency)
        if fx_rate is None:
            stats.unmatched_db.append(_unmatched_shipped_row(r, ref_no=ref_no))
            continue
        unit_price_base = (unit_price_pay * fx_rate).quantize(Decimal("0.000001"))
        shipping_base = (shipping_pay * fx_rate).quantize(Decimal("0.000001"))
        qty = r.get("warehouse_sku_qty") or 0
        try:
            qty_dec = Decimal(int(qty))
        except (TypeError, ValueError):
            qty_dec = Decimal(0)
        order_total_pay = (unit_price_pay * qty_dec + shipping_pay).quantize(Decimal("0.000001"))
        order_total_base = (unit_price_base * qty_dec + shipping_base).quantize(Decimal("0.000001"))
        try:
            did = int(hit.get("id") or 0)
            if did:
                used_detail_ids.add(did)
        except (TypeError, ValueError):
            pass
        updates.append(
            (
                pay_currency,
                unit_price_pay,
                shipping_pay,
                order_total_pay,
                fx_rate,
                unit_price_base,
                shipping_base,
                order_total_base,
                r["id"],
            )
        )

    stats.matched = len(updates)

    if filter_ref_no is None:
        for raw in detail_unique_rows:
            try:
                rid = int(raw.get("id") or 0)
            except (TypeError, ValueError):
                rid = 0
            if not rid or rid in used_detail_ids:
                continue
            if write_window_start is not None and not _detail_row_in_write_window(
                raw, window_start=write_window_start
            ):
                continue
            ref = str(raw.get("ref_no") or "").strip()
            sk = str(raw.get("sku_key") or "").strip()
            up = raw.get("unit_price_pay")
            if not isinstance(up, Decimal):
                try:
                    up = Decimal(str(up)) if up is not None else Decimal("0")
                except InvalidOperation:
                    up = Decimal("0")
            pcc = str(raw.get("pay_currency") or "EUR")
            fxr = _fx_rate_to_base(fx, pcc) or Decimal("1")
            ubase = (up * fxr).quantize(Decimal("0.000001"))
            stats.unmatched_excel.append(
                UnmatchedExcel(
                    sheet=str(raw.get("excel_sheet") or ""),
                    excel_row=int(raw.get("excel_row") or rid),
                    ref_no=ref,
                    sku_key=sk,
                    pay_currency=pcc,
                    unit_price_pay=up,
                    unit_price_base=ubase,
                    order_time=_coerce_datetime_for_notify(raw.get("order_time")),
                )
            )

    if dry_run:
        _LOG.warn(f"--dry-run：跳过 UPDATE sales_order_shipped（本应更新 {stats.matched} 行）")
        stats.updated = 0
        return
    _LOG.info(f"步骤 2b：准备 UPDATE sales_order_shipped：{stats.matched} 行")
    stats.updated = _update_shipped_rows(conn, updates)
    _LOG.info(f"步骤 2b：UPDATE 完成，影响行数={stats.updated}")


def step2_sync_all(
    conn: Any,
    fx: FxRates,
    stats: ImportStats,
    *,
    dry_run: bool,
    filter_ref_no: str | None,
    write_window_start: datetime | None = None,
) -> None:
    step2a_backfill_detail_from_shipped(conn, stats, dry_run=dry_run, write_window_start=write_window_start)
    step2b_sync_shipped_from_detail(
        conn,
        fx,
        stats,
        dry_run=dry_run,
        filter_ref_no=filter_ref_no,
        write_window_start=write_window_start,
    )


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------


def import_file(
    conn: Any,
    xlsx: Path | None = None,
    *,
    dry_run: bool = False,
    filter_ref_no: str | None = None,
    save_detail_table: bool = True,
    steps: tuple[int, ...] | None = None,
    omit_unit_price_pay: bool = False,
    write_order_days: int | None = None,
) -> ImportStats:
    """
    steps:
      None + save_detail_table True  → (1, 2)
      None + save_detail_table False → (2,)  仅同步（兼容 --temu-no-detail-table）
      显式 steps=(1,) / (2,) / (1, 2) 覆盖 save_detail_table

    omit_unit_price_pay:
      为 True 时，步骤 1 写入 temu_order_detail 前将各行的 unit_price_pay 置为 NULL（仍按单价列做有效/无效解析）。

    write_order_days:
      若为正整数，仅写入时间窗口内的数据：步骤 1 按 Excel 解析的 order_time；
      步骤 2a 按 temu_order_detail.order_time；步骤 2b 按 sales_order_shipped.ship_time。
      缺日期列时使用的占位 order_time（2000-01-01）不参与写入窗口。不传则全量。
    """
    if steps is not None:
        run_steps = steps
    else:
        run_steps = (1, 2) if save_detail_table else (2,)

    stats = ImportStats(
        file_name=(xlsx.name if xlsx is not None else "(仅数据库)"),
        started_at=datetime.now().isoformat(timespec="seconds"),
        steps_label="+".join(str(s) for s in run_steps),
    )

    if filter_ref_no:
        _LOG.warn(f"过滤模式：仅处理 ref_no={filter_ref_no!r}（测试用）")

    write_window_start: datetime | None = None
    if write_order_days is not None:
        if write_order_days < 1:
            raise ValueError("write_order_days 须为正整数")
        write_window_start = _write_window_start(days=write_order_days)
        stats.write_order_days_applied = write_order_days
        _LOG.warn(
            f"近 {write_order_days} 天写入模式：order_time / order_time(2a) / ship_time(2b) ≥ {write_window_start.date()}"
        )

    if 1 in run_steps and xlsx is None:
        raise ValueError("步骤 1 需要传入 Excel 路径 xlsx")

    fx = load_rates()
    src_label = (
        f"json:{fx.json_path}"
        if fx.source == "json" and fx.json_path is not None
        else "default(代码内置)"
    )
    meta = []
    if fx.updated_at:
        meta.append(f"updated_at={fx.updated_at}")
    if fx.updated_by:
        meta.append(f"by={fx.updated_by}")
    meta_str = f"  [{' '.join(meta)}]" if meta else ""
    _LOG.info(f"汇率来源：{src_label}{meta_str}")
    _LOG.info(f"汇率（→EUR）：{format_summary(fx)}")
    for issue in fx.issues:
        _LOG.warn(f"汇率配置告警：{issue}")

    _LOG.info(f"执行步骤：{stats.steps_label}")

    if 1 in run_steps:
        assert xlsx is not None
        step1_import_excel_to_detail(
            conn,
            xlsx,
            fx,
            stats,
            dry_run=dry_run,
            omit_unit_price_pay=omit_unit_price_pay,
            write_window_start=write_window_start,
        )

    if 2 in run_steps:
        step2_sync_all(
            conn,
            fx,
            stats,
            dry_run=dry_run,
            filter_ref_no=filter_ref_no,
            write_window_start=write_window_start,
        )

    stats.ended_at = datetime.now().isoformat(timespec="seconds")
    _LOG.info(
        f"结束：detail UPSERT={stats.detail_rows_saved} detail回写={stats.detail_backfilled} "
        f"shipped 匹配={stats.matched} 更新={stats.updated} "
        f"DB缺价格={len(stats.unmatched_db)} 明细未命中shipped={len(stats.unmatched_excel)}"
        + (
            f" | 近{stats.write_order_days_applied}天：明细跳过窗口外={stats.excel_rows_skipped_write_window}"
            if stats.write_order_days_applied
            else ""
        )
    )
    return stats


# ---------------------------------------------------------------------------
# 邮件
# ---------------------------------------------------------------------------


def _load_env_files() -> None:
    if not load_dotenv:
        return
    here = Path(__file__).resolve().parent
    for env_path in (
        here.parent.parent / ".env",
        here.parent / ".env",
        here.parent.parent.parent / ".env",
    ):
        if env_path.is_file():
            load_dotenv(env_path)
            return


def _env_bool(name: str, default: bool) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    return v.strip().lower() in {"1", "true", "yes", "y", "on"}


def _esc(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def _build_html(stats: ImportStats, dry_run: bool) -> str:
    db_mail, ex_mail, cutoff = _unmatched_lists_for_mail(stats)
    n_db_miss = len(db_mail)
    n_ex_miss = len(ex_mail)
    n_db_total = len(stats.unmatched_db)
    n_ex_total = len(stats.unmatched_excel)
    truncated = (n_db_total != n_db_miss) or (n_ex_total != n_ex_miss)
    if n_db_total == 0 and n_ex_total == 0:
        badge = (
            '<span style="display:inline-block;padding:4px 10px;border-radius:999px;'
            'background:#ecfdf5;color:#065f46;font-weight:700;font-size:12px;">✅ 全部匹配</span>'
        )
    else:
        badge = (
            '<span style="display:inline-block;padding:4px 10px;border-radius:999px;'
            f'background:#fef3c7;color:#92400e;font-weight:700;font-size:12px;">'
            f"⚠️ DB缺价格 {n_db_miss} 条 / 明细无对应shipped {n_ex_miss} 条"
            f"（近{MAIL_NOTIFY_ORDER_DAYS}天，自 {cutoff.date()} 起）"
            "</span>"
        )

    mail_filter_tip = ""
    if truncated and (n_db_total > 0 or n_ex_total > 0):
        mail_filter_tip = (
            f"<div style='margin-top:10px;padding:8px 12px;background:#f0f9ff;color:#1e40af;"
            f"border-left:3px solid #3b82f6;font-size:12.5px;'>"
            f"未匹配全量：<b>{n_db_total}</b> 条（shipped）/ <b>{n_ex_total}</b> 条（明细）；"
            f"下表仅展示最近 <b>{MAIL_NOTIFY_ORDER_DAYS}</b> 天内订单（发货时间 ≥ {cutoff.date()}，"
            f"明细按订单时间 ≥ 该日）。无发货时间或订单时间的行不出现在下表。"
            f"</div>"
        )

    DB_LIMIT = 200
    EX_LIMIT = 100
    db_rows_html = "\n".join(
        f"""<tr>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#6b7280;">{r.id}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;font-family:Menlo,Consolas,monospace;color:#111827;">{_esc(r.order_no or '')}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;font-family:Menlo,Consolas,monospace;color:#111827;">{_esc(r.ref_no or '')}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#374151;">{_esc(r.platform_sku_orig or '')}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#374151;">{_esc(r.warehouse_sku or '')}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;text-align:right;color:#374151;">{r.warehouse_sku_qty or ''}</td>
        </tr>""".strip()
        for r in db_mail[:DB_LIMIT]
    )
    excel_rows_html = "\n".join(
        f"""<tr>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#6b7280;">{_esc(r.sheet)}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;text-align:right;color:#6b7280;">{r.excel_row}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;font-family:Menlo,Consolas,monospace;color:#111827;">{_esc(r.ref_no)}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;color:#374151;">{_esc(r.sku_key)}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;text-align:right;color:#374151;">{r.unit_price_pay} {_esc(r.pay_currency)}</td>
          <td style="padding:6px 10px;border:1px solid #e5e7eb;text-align:right;color:#374151;">{r.unit_price_base}</td>
        </tr>""".strip()
        for r in ex_mail[:EX_LIMIT]
    )
    db_more = (
        f"<div style='color:#9ca3af;font-size:12px;margin-top:6px;'>"
        f"… 还有 {n_db_miss - DB_LIMIT} 条未列出</div>"
        if n_db_miss > DB_LIMIT
        else ""
    )
    excel_more = (
        f"<div style='color:#9ca3af;font-size:12px;margin-top:6px;'>"
        f"… 还有 {n_ex_miss - EX_LIMIT} 条未列出</div>"
        if n_ex_miss > EX_LIMIT
        else ""
    )

    db_section = ""
    if n_db_miss > 0:
        db_section = (
            "<h4 style='margin:18px 0 8px 0;color:#111827;'>① sales_order_shipped 缺 temu_order_detail 价格"
            f"（近{n_db_miss} 条）</h4>"
            "<div style='overflow:auto;'>"
            "<table style='width:100%;border-collapse:collapse;font-size:12.5px;'>"
            "<thead><tr style='background:#f9fafb;color:#374151;'>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>id</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>订单号</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>参考号</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>原平台sku</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>仓库sku</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:right;'>销量</th>"
            "</tr></thead><tbody>" + db_rows_html + "</tbody></table></div>" + db_more
        )

    excel_section = ""
    if n_ex_miss > 0:
        excel_section = (
            "<h4 style='margin:18px 0 8px 0;color:#111827;'>② temu_order_detail 有价格但未命中 shipped"
            f"（近{n_ex_miss} 条）</h4>"
            "<div style='overflow:auto;'>"
            "<table style='width:100%;border-collapse:collapse;font-size:12.5px;'>"
            "<thead><tr style='background:#f9fafb;color:#374151;'>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>sheet</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:right;'>行/id</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>参考号</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:left;'>SKU</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:right;'>单价(原币)</th>"
            "<th style='padding:8px 10px;border:1px solid #e5e7eb;text-align:right;'>单价(EUR)</th>"
            "</tr></thead><tbody>" + excel_rows_html + "</tbody></table></div>" + excel_more
        )

    dry_run_tip = (
        "<div style='margin-top:14px;padding:8px 12px;background:#fef9c3;color:#854d0e;"
        "border-left:3px solid #facc15;font-size:12.5px;'>本次以 --dry-run 运行，未实际写库</div>"
        if dry_run
        else ""
    )

    return f"""\
<!doctype html>
<html><head><meta charset="utf-8"/><title>TEMU 费用</title></head>
<body style="margin:0;padding:0;background:#f6f7fb;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,'PingFang SC','Hiragino Sans GB','Microsoft YaHei',sans-serif;">
  <div style="max-width:960px;margin:0 auto;padding:24px;">
    <div style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;">
      <div style="padding:16px 18px;background:linear-gradient(135deg,#0f766e,#115e59);color:#fff;">
        <div style="font-size:16px;font-weight:700;">TEMU 订单费用（分步）</div>
        <div style="margin-top:6px;font-size:12px;opacity:.85;">rpa-task · 步骤 { _esc(stats.steps_label) }</div>
      </div>
      <div style="padding:16px 18px;">
        {badge}
        {dry_run_tip}
        {mail_filter_tip}
        <table style="margin-top:14px;font-size:13px;border-collapse:collapse;">
          <tr><td style="padding:4px 8px;color:#6b7280;">文件</td><td style="padding:4px 8px;color:#111827;font-family:Menlo,Consolas,monospace;">{_esc(stats.file_name)}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">开始 / 结束</td><td style="padding:4px 8px;color:#111827;">{_esc(stats.started_at)} → {_esc(stats.ended_at)}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">Excel 总行 / 取消 / 无效 / 有效</td>
              <td style="padding:4px 8px;color:#111827;">{stats.excel_total_rows} / {stats.excel_cancelled} / {stats.excel_invalid} / {stats.excel_valid_rows}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">Excel 唯一 (参考号, SKU)</td><td style="padding:4px 8px;color:#111827;">{stats.excel_unique_keys}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">DB semitemu 行（步骤2b）</td><td style="padding:4px 8px;color:#111827;">{stats.db_temu_rows}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">temu_order_detail UPSERT</td><td style="padding:4px 8px;color:#111827;">{stats.detail_rows_saved}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">temu 从 shipped 回写</td><td style="padding:4px 8px;color:#111827;">{stats.detail_backfilled}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">shipped 匹配 / UPDATE</td><td style="padding:4px 8px;color:#065f46;font-weight:600;">{stats.matched} / {stats.updated}</td></tr>
          <tr><td style="padding:4px 8px;color:#6b7280;">未匹配（近{MAIL_NOTIFY_ORDER_DAYS}天 / 全量）</td><td style="padding:4px 8px;color:#b45309;font-weight:600;">{n_db_miss} / {n_ex_miss}  ·  {n_db_total} / {n_ex_total}</td></tr>
        </table>
        {db_section}
        {excel_section}
      </div>
    </div>
    <div style="text-align:center;color:#9ca3af;font-size:11px;margin-top:12px;">Generated by rpa-task / import_temu_fee.py</div>
  </div>
</body></html>
"""


def _build_text(stats: ImportStats) -> str:
    db_mail, ex_mail, co = _unmatched_lists_for_mail(stats)
    n_db, n_ex = len(db_mail), len(ex_mail)
    n_dt, n_xt = len(stats.unmatched_db), len(stats.unmatched_excel)
    lines = [
        "TEMU 订单费用（分步）",
        f"步骤：{stats.steps_label}",
        f"文件：{stats.file_name}",
        f"开始：{stats.started_at}    结束：{stats.ended_at}",
        f"Excel 行：总{stats.excel_total_rows} 取消{stats.excel_cancelled} 无效{stats.excel_invalid} 有效{stats.excel_valid_rows}",
        f"Excel 唯一键：{stats.excel_unique_keys}",
        f"DB semitemu 行：{stats.db_temu_rows}",
        f"temu_order_detail UPSERT：{stats.detail_rows_saved}",
        f"temu 从 shipped 回写：{stats.detail_backfilled}",
        f"shipped 匹配：{stats.matched}    UPDATE：{stats.updated}",
        f"未匹配 全量：DB {n_dt} / 明细 {n_xt}；近{MAIL_NOTIFY_ORDER_DAYS}天（自{co.date()}）：DB {n_db} / 明细 {n_ex}",
    ]
    if n_dt or n_xt:
        lines.append(
            f"（下列清单仅列近 {MAIL_NOTIFY_ORDER_DAYS} 天内、且具备发货时间/订单时间的行）"
        )
    if db_mail:
        lines.append("")
        lines.append("=== ① sales_order_shipped 缺明细价格（时间窗口内）===")
        lines.append("id\torder_no\tref_no\t原平台sku\t仓库sku\t销量")
        for r in db_mail[:200]:
            lines.append(
                f"{r.id}\t{r.order_no or ''}\t{r.ref_no or ''}\t"
                f"{r.platform_sku_orig or ''}\t{r.warehouse_sku or ''}\t{r.warehouse_sku_qty or ''}"
            )
    if ex_mail:
        lines.append("")
        lines.append("=== ② temu_order_detail 未命中 shipped（时间窗口内）===")
        lines.append("sheet\trow\tref_no\tsku\tunit_price_pay\tpay_currency\tunit_price_base(EUR)")
        for r in ex_mail[:100]:
            lines.append(
                f"{r.sheet}\t{r.excel_row}\t{r.ref_no}\t{r.sku_key}\t"
                f"{r.unit_price_pay}\t{r.pay_currency}\t{r.unit_price_base}"
            )
    return "\n".join(lines) + "\n"


def send_notification(stats: ImportStats, *, dry_run: bool) -> bool:
    host = os.getenv("SMTP_HOST", "").strip()
    port_raw = os.getenv("SMTP_PORT", "587").strip()
    user = os.getenv("SMTP_USER") or None
    pwd = os.getenv("SMTP_PASS") or None
    use_starttls = _env_bool("SMTP_STARTTLS", True)
    use_ssl = _env_bool("SMTP_SSL", False)
    mail_from = os.getenv("MAIL_FROM", "").strip() or (user or "")
    mail_to_raw = os.getenv("MAIL_TO", "").strip()
    subject_prefix = os.getenv("MAIL_SUBJECT", "").strip() or "rpa-task"

    if not host or not mail_from or not mail_to_raw:
        _LOG.warn("缺少 SMTP/邮件配置（SMTP_HOST/MAIL_FROM/MAIL_TO），跳过邮件通知")
        return False

    try:
        port = int(port_raw)
    except ValueError:
        port = 587

    mail_to = [s.strip() for s in mail_to_raw.replace(";", ",").split(",") if s.strip()]
    if not mail_to:
        _LOG.warn("MAIL_TO 为空，跳过邮件通知")
        return False

    db_mail, ex_mail, _ = _unmatched_lists_for_mail(stats)
    n_db = len(db_mail)
    n_ex = len(ex_mail)
    n_db_all = len(stats.unmatched_db)
    n_ex_all = len(stats.unmatched_excel)
    if n_db_all == 0 and n_ex_all == 0:
        subject = f"[{subject_prefix}] TEMU 费用 - 全部匹配（{stats.matched} 行）"
    elif n_db == 0 and n_ex == 0:
        subject = (
            f"[{subject_prefix}] TEMU 费用 - 近{MAIL_NOTIFY_ORDER_DAYS}天无未匹配项"
            f"（全量 DB缺{n_db_all} / 明细{n_ex_all}，见正文说明）"
        )
    else:
        subject = (
            f"[{subject_prefix}] TEMU 费用 - 近{MAIL_NOTIFY_ORDER_DAYS}天 DB缺价 {n_db} / 明细无shipped {n_ex}"
        )
        if n_db != n_db_all or n_ex != n_ex_all:
            subject += f"（全量 {n_db_all}/{n_ex_all}）"

    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = ", ".join(mail_to)
    msg["Subject"] = subject
    msg.set_content(_build_text(stats))
    msg.add_alternative(_build_html(stats, dry_run), subtype="html")

    try:
        if use_ssl:
            with smtplib.SMTP_SSL(host, port, timeout=30) as server:
                if user and pwd:
                    server.login(user, pwd)
                server.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=30) as server:
                server.ehlo()
                if use_starttls:
                    server.starttls()
                    server.ehlo()
                if user and pwd:
                    server.login(user, pwd)
                server.send_message(msg)
        _LOG.info(f"邮件已发送：to={mail_to} subject={subject!r}")
        return True
    except Exception as e:
        _LOG.error(f"邮件发送失败：{type(e).__name__}: {e}")
        return False


def default_temu_excel_path() -> Path:
    return default_order_excel_dir() / DEFAULT_FILE_NAME


def main() -> int:
    setup_stdout_utf8()
    _load_env_files()

    ap = argparse.ArgumentParser(
        description="TEMU：步骤1 导入 Excel→temu_order_detail；步骤2 从明细补全并发货表费用"
    )
    ap.add_argument("--file", type=Path, default=None, help=f"xlsx 路径，默认 {default_temu_excel_path()}")
    ap.add_argument("--dry-run", action="store_true", help="不写库，仅演练")
    ap.add_argument("--no-mail", action="store_true")
    ap.add_argument("--always-mail", action="store_true")
    ap.add_argument("--ref-no", type=str, default=None, metavar="PO-XXX", help="仅处理某参考号（测试）")
    step = ap.add_mutually_exclusive_group()
    step.add_argument("--only-step1", action="store_true", help="只执行步骤 1（必须能读取 Excel）")
    step.add_argument("--only-step2", action="store_true", help="只执行步骤 2（仅数据库，可不提供 Excel）")
    ap.add_argument(
        "--no-detail-table",
        action="store_true",
        help="兼容旧参数：等同 --only-step2（不写明细、只同步 shipped）",
    )
    ap.add_argument(
        "--write-order-days",
        type=int,
        default=None,
        metavar="N",
        help="仅写入近 N 天：步骤1 按明细 order_time；步骤2a 按 temu_order_detail.order_time；"
        "步骤2b 按 sales_order_shipped.ship_time。缺日期列占位 2000-01-01 的行不写入。不传则全量。",
    )
    args = ap.parse_args()

    if args.write_order_days is not None and args.write_order_days < 1:
        ap.error("--write-order-days 须为 >=1 的整数")

    if args.no_detail_table and args.only_step1:
        ap.error("--no-detail-table 与 --only-step1 互斥")
    if args.only_step1:
        steps: tuple[int, ...] = (1,)
    elif args.only_step2 or args.no_detail_table:
        steps = (2,)
    else:
        steps = (1, 2)

    if 1 in steps:
        xlsx = args.file or default_temu_excel_path()
        if not xlsx.is_file():
            _LOG.error(f"Excel 不存在：{xlsx}")
            return 2
    else:
        xlsx = args.file

    cfg = load_db_config()
    _LOG.info(f"连接数据库：host={cfg.host} port={cfg.port} database={cfg.database} user={cfg.user}")
    conn = connect(cfg)
    try:
        stats = import_file(
            conn,
            xlsx,
            dry_run=args.dry_run,
            filter_ref_no=args.ref_no,
            steps=steps,
            write_order_days=args.write_order_days,
        )
        if args.dry_run:
            conn.rollback()
            _LOG.warn("--dry-run：已回滚事务")
        else:
            conn.commit()
            _LOG.info(f"已提交事务：shipped UPDATE 影响行={stats.updated}")
    except Exception:
        conn.rollback()
        _LOG.error("发生异常，已回滚事务")
        raise
    finally:
        conn.close()
        _LOG.info("数据库连接已关闭")

    n_unmatched = len(stats.unmatched_db) + len(stats.unmatched_excel)
    if args.no_mail:
        _LOG.info("--no-mail 已设置，跳过邮件通知")
    elif args.always_mail or n_unmatched > 0:
        send_notification(stats, dry_run=args.dry_run)
    else:
        _LOG.info("无未匹配项，按默认策略不发邮件（如需总是发送请加 --always-mail）")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
